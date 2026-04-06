"""
escalation_mailer.py
====================
Builds and sends two-tier escalation emails with an Excel attachment
summarising the ticket.

Dependencies
------------
    pip install openpyxl   (already likely installed for bulk-upload)

Settings used
-------------
    EMAIL_HOST, EMAIL_PORT, EMAIL_HOST_USER, EMAIL_HOST_PASSWORD,
    EMAIL_USE_TLS / EMAIL_USE_SSL   — standard Django email settings
    DEFAULT_FROM_EMAIL              — sender address
    KSTS_SYSTEM_NAME                — optional, defaults to "KSTS"
    KSTS_BASE_URL                   — used for "View Ticket" link in email
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Optional

from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone

logger = logging.getLogger("ksts.escalation_mailer")

SYSTEM_NAME: str = getattr(settings, "KSTS_SYSTEM_NAME", "KSTS")
BASE_URL: str = getattr(settings, "KSTS_BASE_URL", "http://localhost:8000")


# ═══════════════════════════════════════════════════════════════════════════════
# PUBLIC CLASS
# ═══════════════════════════════════════════════════════════════════════════════

class EscalationMailer:
    """
    Sends tier-1 (manager) and tier-2 (C.E.) escalation emails.
    Each email carries an Excel attachment with full ticket details.
    """

    # ── Tier-1 ──────────────────────────────────────────────────────────────
    def send_tier1(self, *, ticket, recipient_email: str) -> None:
        subject = (
            f"[{SYSTEM_NAME}] ⚠️ Escalated Ticket Awaiting Action — {ticket.ticket_id}"
        )
        html_body = self._build_tier1_html(ticket)
        excel_bytes = self._build_excel(ticket, tier=1)
        self._dispatch(
            subject=subject,
            html_body=html_body,
            recipient=recipient_email,
            ticket=ticket,
            excel_bytes=excel_bytes,
            filename=f"Escalated_{ticket.ticket_id}_Manager_Report.xlsx",
        )

    # ── Tier-2 ──────────────────────────────────────────────────────────────
    def send_tier2(self, *, ticket, recipient_email: str) -> None:
        subject = (
            f"[{SYSTEM_NAME}] 🚨 URGENT — Unresolved Escalation Requires CE Attention — {ticket.ticket_id}"
        )
        html_body = self._build_tier2_html(ticket)
        excel_bytes = self._build_excel(ticket, tier=2)
        self._dispatch(
            subject=subject,
            html_body=html_body,
            recipient=recipient_email,
            ticket=ticket,
            excel_bytes=excel_bytes,
            filename=f"Escalated_{ticket.ticket_id}_CE_Report.xlsx",
        )

    # ═══════════════════════════════════════════════════════════════════════════
    # EMAIL BUILDERS
    # ═══════════════════════════════════════════════════════════════════════════

    def _build_tier1_html(self, ticket) -> str:
        manager_name = "Reporting Manager"
        if ticket.escalated_to and ticket.escalated_to.manager:
            manager_name = ticket.escalated_to.manager.get_full_name() or manager_name

        escalated_by_name = (
            ticket.escalated_to.get_full_name() if ticket.escalated_to else "System"
        )
        hours_open = self._hours_since(ticket.escalated_at or ticket.updated_at)

        return _render_email_html(
            tier=1,
            badge_text="Manager Action Required",
            badge_color="#E65100",
            header_text="Ticket Escalation — Manager Notification",
            greeting=f"Dear {manager_name},",
            intro=(
                f"A support ticket has been <strong>escalated</strong> and is awaiting your "
                f"intervention. The ticket has been open for <strong>{hours_open:.0f} hours</strong> "
                f"without resolution."
            ),
            ticket=ticket,
            cta_text="Review & Act on Ticket",
            footer_note=(
                "If this ticket remains unresolved, it will be automatically escalated "
                f"to the Chief Executive after {getattr(settings, 'ESCALATION_TIER2_GRACE_HOURS', 24)} hours."
            ),
            escalated_by=escalated_by_name,
        )

    def _build_tier2_html(self, ticket) -> str:
        from_settings = getattr(settings, "ESCALATION_CE_NAME", "Chief Executive")
        hours_open = self._hours_since(ticket.escalated_at or ticket.updated_at)

        manager_name = "—"
        if ticket.escalated_to and ticket.escalated_to.manager:
            manager_name = ticket.escalated_to.manager.get_full_name()

        escalated_by_name = (
            ticket.escalated_to.get_full_name() if ticket.escalated_to else "System"
        )

        return _render_email_html(
            tier=2,
            badge_text="CE Escalation — Urgent",
            badge_color="#B71C1C",
            header_text="Critical Escalation — Chief Executive Notification",
            greeting=f"Dear {from_settings},",
            intro=(
                f"A support ticket has been escalated to your attention. "
                f"It has been open for <strong>{hours_open:.0f} hours</strong> without resolution. "
                f"The reporting manager (<strong>{manager_name}</strong>) has already been notified "
                f"but the issue remains unresolved."
            ),
            ticket=ticket,
            cta_text="Review Escalated Ticket",
            footer_note=(
                "This is an automated tier-2 escalation alert from the KSTS system. "
                "No further automatic escalations will occur for this ticket."
            ),
            escalated_by=escalated_by_name,
        )

    # ═══════════════════════════════════════════════════════════════════════════
    # EXCEL BUILDER
    # ═══════════════════════════════════════════════════════════════════════════

    def _build_excel(self, ticket, tier: int) -> bytes:
        """
        Build an in-memory Excel workbook with full ticket details.
        Returns raw bytes ready to attach to the email.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side, GradientFill
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed — Excel attachment skipped.")
            return b""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Escalation Report"

        # ── Colour palette ────────────────────────────────────────────────
        RED_DARK   = "B71C1C"
        RED_MID    = "E53935"
        ORANGE     = "E65100"
        AMBER      = "FF8F00"
        CREAM      = "FFF8E1"
        WHITE      = "FFFFFF"
        GREY_LIGHT = "F5F5F5"
        GREY_MID   = "BDBDBD"
        DARK       = "212121"

        header_fill_hex = RED_DARK if tier == 2 else ORANGE

        thin = Side(style="thin", color=GREY_MID)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Helper lambdas ────────────────────────────────────────────────
        def hdr_cell(row, col, value, bg=header_fill_hex, fg=WHITE, bold=True, size=11):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
            return c

        def data_cell(row, col, value, bg=WHITE, bold=False, align="left", color=DARK):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=bold, color=color, size=10, name="Calibri")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            c.border = border
            return c

        def section_header(row, col_start, col_end, text, bg=AMBER):
            ws.merge_cells(
                start_row=row, start_column=col_start,
                end_row=row, end_column=col_end,
            )
            c = ws.cell(row=row, column=col_start, value=text)
            c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border

        # ── Title row ─────────────────────────────────────────────────────
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        tier_label = "TIER-2 (CE)" if tier == 2 else "TIER-1 (MANAGER)"
        title_cell.value = f"KSTS ESCALATION REPORT — {tier_label}"
        title_cell.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
        title_cell.fill = PatternFill("solid", fgColor=header_fill_hex)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # Generated timestamp
        ws.merge_cells("A2:F2")
        ts_cell = ws["A2"]
        ts_cell.value = f"Generated: {timezone.now().strftime('%d %b %Y, %I:%M %p')}"
        ts_cell.font = Font(italic=True, color="757575", size=9, name="Calibri")
        ts_cell.fill = PatternFill("solid", fgColor=CREAM)
        ts_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[2].height = 18

        # ── Section 1: Ticket Identity ─────────────────────────────────────
        section_header(3, 1, 6, "  📋  TICKET IDENTITY", RED_MID)
        ws.row_dimensions[3].height = 20

        ticket_rows = [
            ("Ticket ID",      ticket.ticket_id,                         True,  CREAM),
            ("Status",         ticket.get_status_display(),               True,  CREAM),
            ("Priority",       ticket.get_priority_display(),             True,  CREAM),
            ("Type",           ticket.get_ticket_type_display(),          False, WHITE),
            ("Entity Type",    ticket.get_entity_type_display(),          False, WHITE),
            ("Created At",     _fmt_dt(ticket.created_at),               False, WHITE),
            ("Escalated At",   _fmt_dt(ticket.escalated_at),             True,  CREAM),
            ("Expected Res.",  _fmt_date(ticket.expected_resolution_date),False, WHITE),
        ]
        r = 4
        hdr_cell(r - 1 + 1 - 1, 1, "Field",  bg="424242"); hdr_cell(r - 1 + 1 - 1, 2, "Value", bg="424242")  # noqa — we'll do it properly below
        # Column headers for ticket identity
        hdr_cell(4, 1, "Field",  bg="424242", size=10)
        hdr_cell(4, 2, "Value",  bg="424242", size=10)
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)
        r = 5
        for label, value, bold, bg in ticket_rows:
            data_cell(r, 1, label, bg=GREY_LIGHT, bold=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            data_cell(r, 2, value, bg=bg, bold=bold)
            r += 1

        # ── Section 2: Escalation Details ─────────────────────────────────
        section_header(r, 1, 6, "  🚨  ESCALATION DETAILS", header_fill_hex)
        ws.row_dimensions[r].height = 20
        r += 1

        escalated_to_name = (
            ticket.escalated_to.get_full_name() if ticket.escalated_to else "—"
        )
        escalated_to_email = (
            ticket.escalated_to.email if ticket.escalated_to else "—"
        )
        manager_name = "—"
        manager_email = "—"
        if ticket.escalated_to and ticket.escalated_to.manager:
            mgr = ticket.escalated_to.manager
            manager_name  = mgr.get_full_name() or "—"
            manager_email = mgr.email or "—"

        esc_rows = [
            ("Escalated To",          escalated_to_name),
            ("Escalated To Email",     escalated_to_email),
            ("Reporting Manager",      manager_name),
            ("Manager Email",          manager_email),
            ("Escalation Reason",      ticket.escalation_reason or "No reason provided"),
            ("Hours Since Escalation", f"{self._hours_since(ticket.escalated_at or ticket.updated_at):.1f} hrs"),
        ]
        for label, value in esc_rows:
            data_cell(r, 1, label, bg=GREY_LIGHT, bold=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            data_cell(r, 2, value, bg=WHITE)
            r += 1

        # ── Section 3: Caller / Entity ─────────────────────────────────────
        section_header(r, 1, 6, "  👤  CALLER / ENTITY", "37474F")
        ws.row_dimensions[r].height = 20
        r += 1

        caller_rows = [
            ("Caller Name",     ticket.caller_display_name),
            ("Caller Mobile",   ticket.caller_contact_mobile or "—"),
            ("Caller Relation", ticket.get_caller_relation_display() if ticket.caller_relation else "—"),
            ("Location",        ticket.caller_location),
        ]
        for label, value in caller_rows:
            data_cell(r, 1, label, bg=GREY_LIGHT, bold=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            data_cell(r, 2, value, bg=WHITE)
            r += 1

        # ── Section 4: Issue Description ──────────────────────────────────
        section_header(r, 1, 6, "  📝  ISSUE DESCRIPTION", "4A148C")
        ws.row_dimensions[r].height = 20
        r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        desc_cell = ws.cell(
            row=r, column=1,
            value=ticket.description_en or "(No description provided)",
        )
        desc_cell.font = Font(size=10, name="Calibri", color=DARK)
        desc_cell.fill = PatternFill("solid", fgColor=CREAM)
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        desc_cell.border = border
        ws.row_dimensions[r].height = 60
        r += 1

        # ── Section 5: Assignment ─────────────────────────────────────────
        section_header(r, 1, 6, "  👥  ASSIGNED EMPLOYEES", "1B5E20")
        ws.row_dimensions[r].height = 20
        r += 1

        hdr_cell(r, 1, "Name",       bg="2E7D32", size=9)
        hdr_cell(r, 2, "Emp. Code",  bg="2E7D32", size=9)
        hdr_cell(r, 3, "Title",      bg="2E7D32", size=9)
        hdr_cell(r, 4, "Department", bg="2E7D32", size=9)
        hdr_cell(r, 5, "Email",      bg="2E7D32", size=9)
        hdr_cell(r, 6, "Mobile",     bg="2E7D32", size=9)
        r += 1

        assignees = list(ticket.assigned_to.all())
        if assignees:
            for emp in assignees:
                bg = GREY_LIGHT if r % 2 == 0 else WHITE
                data_cell(r, 1, emp.get_full_name(), bg=bg)
                data_cell(r, 2, emp.employee_code,   bg=bg)
                data_cell(r, 3, emp.employee_title or "—", bg=bg)
                data_cell(r, 4, emp.department,      bg=bg)
                data_cell(r, 5, emp.email,            bg=bg)
                data_cell(r, 6, emp.mobile_number or "—", bg=bg)
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            data_cell(r, 1, "No employees assigned", bg=CREAM, color="757575")
            r += 1

        # ── Section 6: Recent Comments ────────────────────────────────────
        section_header(r, 1, 6, "  💬  RECENT COMMENTS (last 5)", "004D40")
        ws.row_dimensions[r].height = 20
        r += 1

        hdr_cell(r, 1, "Date/Time",  bg="00695C", size=9)
        hdr_cell(r, 2, "Posted By",  bg="00695C", size=9)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        hdr_cell(r, 3, "Comment",    bg="00695C", size=9)
        r += 1

        recent_comments = ticket.comments.order_by("-created_at")[:5]
        if recent_comments:
            for comment in recent_comments:
                bg = GREY_LIGHT if r % 2 == 0 else WHITE
                data_cell(r, 1, _fmt_dt(comment.created_at), bg=bg)
                poster = comment.posted_by.get_full_name() if comment.posted_by else "System"
                data_cell(r, 2, poster, bg=bg)
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
                data_cell(r, 3, comment.body_text[:300] or "(empty)", bg=bg)
                ws.row_dimensions[r].height = 40
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            data_cell(r, 1, "No comments yet", bg=CREAM, color="757575")
            r += 1

        # ── Column widths ─────────────────────────────────────────────────
        col_widths = {1: 28, 2: 22, 3: 22, 4: 22, 5: 30, 6: 18}
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # ── Freeze top rows ───────────────────────────────────────────────
        ws.freeze_panes = "A3"

        # ── Write to bytes ────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ═══════════════════════════════════════════════════════════════════════════
    # DISPATCH
    # ═══════════════════════════════════════════════════════════════════════════

    def _dispatch(
        self,
        *,
        subject: str,
        html_body: str,
        recipient: str,
        ticket,
        excel_bytes: bytes,
        filename: str,
    ) -> None:
        from_email = settings.DEFAULT_FROM_EMAIL
        msg = EmailMessage(
            subject=subject,
            body=html_body,
            from_email=from_email,
            to=[recipient],
        )
        msg.content_subtype = "html"

        if excel_bytes:
            msg.attach(
                filename=filename,
                content=excel_bytes,
                mimetype=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
            )

        try:
            msg.send(fail_silently=False)
            logger.info("Escalation mail dispatched | to=%s subject=%s", recipient, subject)
        except Exception as exc:
            logger.exception(
                "Failed to send escalation mail | to=%s ticket=%s | %s",
                recipient, ticket.ticket_id, exc,
            )
            raise

    # ── Utility ───────────────────────────────────────────────────────────────
    @staticmethod
    def _hours_since(dt) -> float:
        if not dt:
            return 0.0
        return (timezone.now() - dt).total_seconds() / 3600


# ═══════════════════════════════════════════════════════════════════════════════
# HTML EMAIL TEMPLATE
# ═══════════════════════════════════════════════════════════════════════════════

def _fmt_dt(dt) -> str:
    if not dt:
        return "—"
    local = timezone.localtime(dt)
    return local.strftime("%d %b %Y, %I:%M %p")


def _fmt_date(d) -> str:
    if not d:
        return "—"
    if isinstance(d, datetime):
        return d.strftime("%d %b %Y")
    return d.strftime("%d %b %Y")


def _render_email_html(
    *,
    tier: int,
    badge_text: str,
    badge_color: str,
    header_text: str,
    greeting: str,
    intro: str,
    ticket,
    cta_text: str,
    footer_note: str,
    escalated_by: str,
) -> str:
    ticket_url = f"{BASE_URL}/my_tickets/?ticket={ticket.ticket_id}"

    priority_colors = {
        "low":      ("#1B5E20", "#E8F5E9"),
        "medium":   ("#E65100", "#FFF3E0"),
        "high":     ("#B71C1C", "#FFEBEE"),
        "critical": ("#4A148C", "#F3E5F5"),
    }
    p_color, p_bg = priority_colors.get(ticket.priority, ("#424242", "#F5F5F5"))

    status_colors = {
        "escalated": ("#B71C1C", "#FFEBEE"),
        "open":      ("#1565C0", "#E3F2FD"),
        "pending":   ("#E65100", "#FFF3E0"),
        "reopened":  ("#4A148C", "#F3E5F5"),
    }
    s_color, s_bg = status_colors.get(ticket.status, ("#424242", "#F5F5F5"))

    caller_name    = ticket.caller_display_name
    caller_mobile  = ticket.caller_contact_mobile or "—"
    caller_loc     = ticket.caller_location
    description    = ticket.description_en or "(No description provided)"
    escalation_rsn = ticket.escalation_reason or "Not specified"

    assignees_html = ""
    for emp in ticket.assigned_to.all():
        assignees_html += f"""
        <tr>
          <td style="padding:8px 12px;border-bottom:1px solid #eee;">
            {emp.get_full_name()}
          </td>
          <td style="padding:8px 12px;border-bottom:1px solid #eee;color:#666;">
            {emp.employee_title or emp.department}
          </td>
          <td style="padding:8px 12px;border-bottom:1px solid #eee;color:#666;">
            {emp.email}
          </td>
        </tr>"""
    if not assignees_html:
        assignees_html = """
        <tr>
          <td colspan="3" style="padding:8px 12px;color:#999;font-style:italic;">
            No employees assigned
          </td>
        </tr>"""

    tier2_warning = ""
    if tier == 1:
        grace2 = getattr(settings, "ESCALATION_TIER2_GRACE_HOURS", 24)
        tier2_warning = f"""
        <div style="
            margin:20px 0;
            padding:14px 18px;
            background:#FFF3E0;
            border-left:4px solid #E65100;
            border-radius:4px;
            font-size:13px;
            color:#BF360C;
        ">
            ⏰ <strong>Note:</strong> If this ticket is not resolved within
            {grace2} hours, it will be automatically escalated to the
            <strong>Chief Executive</strong>.
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{header_text}</title>
</head>
<body style="margin:0;padding:0;background:#F0F2F5;font-family:'Segoe UI',Arial,sans-serif;">

<table width="100%" cellpadding="0" cellspacing="0" style="background:#F0F2F5;padding:32px 16px;">
<tr><td align="center">
<table width="620" cellpadding="0" cellspacing="0" style="
    background:#FFFFFF;
    border-radius:8px;
    overflow:hidden;
    box-shadow:0 4px 24px rgba(0,0,0,0.10);
">

  <!-- ── HEADER BAND ── -->
  <tr>
    <td style="background:{badge_color};padding:28px 32px;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td>
            <div style="
                display:inline-block;
                background:rgba(255,255,255,0.18);
                color:#fff;
                font-size:11px;
                font-weight:700;
                letter-spacing:1.5px;
                text-transform:uppercase;
                padding:4px 12px;
                border-radius:20px;
                margin-bottom:10px;
            ">{badge_text}</div>
            <h1 style="
                margin:0;
                color:#FFFFFF;
                font-size:22px;
                font-weight:700;
                line-height:1.3;
            ">{header_text}</h1>
            <p style="margin:6px 0 0;color:rgba(255,255,255,0.80);font-size:13px;">
              {SYSTEM_NAME} · Ticket Management System
            </p>
          </td>
          <td align="right" style="vertical-align:top;">
            <div style="
                background:rgba(255,255,255,0.15);
                border-radius:6px;
                padding:10px 16px;
                text-align:center;
                min-width:120px;
            ">
              <div style="color:rgba(255,255,255,0.75);font-size:11px;letter-spacing:1px;
                          text-transform:uppercase;margin-bottom:4px;">Ticket ID</div>
              <div style="color:#FFFFFF;font-size:18px;font-weight:800;
                          font-family:monospace;">{ticket.ticket_id}</div>
            </div>
          </td>
        </tr>
      </table>
    </td>
  </tr>

  <!-- ── BODY ── -->
  <tr>
    <td style="padding:32px;">

      <!-- Greeting + intro -->
      <p style="margin:0 0 8px;font-size:15px;color:#212121;font-weight:600;">{greeting}</p>
      <p style="margin:0 0 20px;font-size:14px;color:#424242;line-height:1.7;">{intro}</p>

      {tier2_warning}

      <!-- Status / Priority pills -->
      <table cellpadding="0" cellspacing="0" style="margin-bottom:24px;">
        <tr>
          <td style="padding-right:12px;">
            <span style="
                background:{s_bg};color:{s_color};
                padding:4px 14px;border-radius:20px;
                font-size:12px;font-weight:700;
                border:1px solid {s_color};
                text-transform:uppercase;letter-spacing:0.5px;
            ">{ticket.get_status_display()}</span>
          </td>
          <td>
            <span style="
                background:{p_bg};color:{p_color};
                padding:4px 14px;border-radius:20px;
                font-size:12px;font-weight:700;
                border:1px solid {p_color};
                text-transform:uppercase;letter-spacing:0.5px;
            ">{ticket.get_priority_display()} Priority</span>
          </td>
        </tr>
      </table>

      <!-- ── Ticket Details Card ── -->
      <div style="
          background:#FAFAFA;
          border:1px solid #E0E0E0;
          border-radius:8px;
          overflow:hidden;
          margin-bottom:24px;
      ">
        <div style="
            background:#37474F;
            padding:10px 16px;
            font-size:12px;
            font-weight:700;
            color:#FFFFFF;
            letter-spacing:0.8px;
            text-transform:uppercase;
        ">📋 Ticket Details</div>

        <table width="100%" cellpadding="0" cellspacing="0">
          {''.join(_detail_row(k, v) for k, v in [
              ("Ticket Type",     ticket.get_ticket_type_display()),
              ("Entity",          ticket.get_entity_type_display()),
              ("Caller Name",     caller_name),
              ("Caller Mobile",   caller_mobile),
              ("Location",        caller_loc),
              ("Created At",      _fmt_dt(ticket.created_at)),
              ("Escalated At",    _fmt_dt(ticket.escalated_at)),
              ("Escalated By",    escalated_by),
              ("Escalation Reason", escalation_rsn),
              ("Expected Resolution", _fmt_date(ticket.expected_resolution_date)),
          ])}
        </table>
      </div>

      <!-- ── Issue Description ── -->
      <div style="
          background:#FFF8E1;
          border:1px solid #FFE082;
          border-radius:8px;
          overflow:hidden;
          margin-bottom:24px;
      ">
        <div style="
            background:#F57F17;
            padding:10px 16px;
            font-size:12px;
            font-weight:700;
            color:#FFFFFF;
            letter-spacing:0.8px;
            text-transform:uppercase;
        ">📝 Issue Description</div>
        <div style="padding:16px;font-size:13px;color:#424242;line-height:1.7;
                    white-space:pre-wrap;">{description}</div>
      </div>

      <!-- ── Assigned Employees ── -->
      <div style="
          border:1px solid #E0E0E0;
          border-radius:8px;
          overflow:hidden;
          margin-bottom:24px;
      ">
        <div style="
            background:#2E7D32;
            padding:10px 16px;
            font-size:12px;
            font-weight:700;
            color:#FFFFFF;
            letter-spacing:0.8px;
            text-transform:uppercase;
        ">👥 Assigned Employees</div>
        <table width="100%" cellpadding="0" cellspacing="0">
          <tr style="background:#F5F5F5;">
            <th style="padding:8px 12px;text-align:left;font-size:11px;
                       color:#666;border-bottom:1px solid #eee;">Name</th>
            <th style="padding:8px 12px;text-align:left;font-size:11px;
                       color:#666;border-bottom:1px solid #eee;">Role</th>
            <th style="padding:8px 12px;text-align:left;font-size:11px;
                       color:#666;border-bottom:1px solid #eee;">Email</th>
          </tr>
          {assignees_html}
        </table>
      </div>

      <!-- ── CTA Button ── -->
      <div style="text-align:center;margin:28px 0;">
        <a href="{ticket_url}" style="
            display:inline-block;
            background:{badge_color};
            color:#FFFFFF;
            text-decoration:none;
            padding:14px 36px;
            border-radius:6px;
            font-size:15px;
            font-weight:700;
            letter-spacing:0.3px;
            box-shadow:0 4px 12px rgba(0,0,0,0.20);
        ">{cta_text} →</a>
      </div>

      <!-- Footer note -->
      <p style="
          margin:0;
          padding:16px;
          background:#F5F5F5;
          border-radius:6px;
          font-size:12px;
          color:#757575;
          line-height:1.6;
          border-left:3px solid #BDBDBD;
      ">{footer_note}</p>

    </td>
  </tr>

  <!-- ── FOOTER ── -->
  <tr>
    <td style="
        background:#37474F;
        padding:18px 32px;
        text-align:center;
    ">
      <p style="margin:0;font-size:11px;color:rgba(255,255,255,0.55);">
        This is an automated notification from <strong style="color:rgba(255,255,255,0.80);">
        {SYSTEM_NAME}</strong>. Please do not reply to this email.
      </p>
      <p style="margin:6px 0 0;font-size:11px;color:rgba(255,255,255,0.40);">
        Sent: {timezone.now().strftime("%d %b %Y, %I:%M %p")}
      </p>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>"""


def _detail_row(label: str, value: str) -> str:
    return f"""
    <tr>
      <td style="
          padding:10px 16px;
          font-size:12px;
          font-weight:600;
          color:#616161;
          background:#F9F9F9;
          width:38%;
          border-bottom:1px solid #EEEEEE;
          vertical-align:top;
      ">{label}</td>
      <td style="
          padding:10px 16px;
          font-size:13px;
          color:#212121;
          border-bottom:1px solid #EEEEEE;
      ">{value or '—'}</td>
    </tr>"""