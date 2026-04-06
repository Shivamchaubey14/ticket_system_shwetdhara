# main_app_ticket/escalation_engine.py
"""
KSTS Production Escalation Engine  v2
======================================

TWO TRIGGERS for escalation emails:

  TRIGGER A — Manual escalation (existing button in dashboard):
      ticket.escalate() called in api_views
          → Tier-1 fires immediately  → managers of all assignees
          → Tier-2 fires TIER2_DELAY_HOURS later if still unresolved → C.E.

  TRIGGER B — Automatic overdue detection (THE FIX for the reported bug):
      Celery Beat runs `overdue_auto_escalate_sweep` every 15 minutes.
      Any ticket that is:
        • status in (open, pending, reopened)
        • past its expected_resolution_date
        • NOT already escalated / resolved / closed
      … is automatically:
        1. Marked escalated in the DB
        2. Logged as a TicketActivity (performed_by=None = system)
        3. Tier-1 email + Excel fired immediately to assignees' managers
        4. Tier-2 task scheduled for TIER2_DELAY_HOURS later

  TIER-2 SAFETY NET:
      `tier2_sweep` runs every 30 min and catches any tickets where
      Tier-1 was sent TIER2_DELAY_HOURS+ ago but Tier-2 hasn't fired.
"""

import io
import logging
import re
from collections import Counter
from datetime import timedelta

from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

logger = logging.getLogger("ksts.escalation")

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

TIER2_DELAY_HOURS = getattr(settings, "ESCALATION_TIER2_DELAY_HOURS", 24)
CE_EMAIL          = getattr(settings, "ESCALATION_CE_EMAIL", None)
SENDER_NAME       = getattr(settings, "ESCALATION_SENDER_NAME", "Shwetdhara Dairy — KSTS")
SITE_URL          = getattr(settings, "SITE_URL", "http://localhost:8000")

_ACTIVE_STATUSES  = ("open", "pending", "reopened")


# ─────────────────────────────────────────────────────────────────────────────
#  STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

_G_DARK = "3A6B4A"; _G_MID = "4E7D5E"; _G_LIGHT = "C5D9C0"
_NAVY   = "2C3547"; _SLATE  = "4A5568"; _ALT     = "F2F6F1"
_WHITE  = "FFFFFF"; _CREAM  = "FFF8DC"; _GOLD    = "7B5700"
_RED_BG = "FEE2E2"; _RED_FG = "991B1B"
_AMB_BG = "FEF9C3"; _AMB_FG = "854D0E"

_PRIO_C = {
    "low":      ("DCFCE7", "166534"),
    "medium":   ("FEF9C3", "854D0E"),
    "high":     ("FFEDD5", "9A3412"),
    "critical": ("FEE2E2", "991B1B"),
}
_ACT_DOT = {
    "created": "2C3547", "assigned": "0A7A8E", "reassigned": "435663",
    "comment": "6F42C1", "attachment": "435663", "resolved": "1E7E34",
    "closed": "5A6270",  "reopened": "FD7E14",   "escalated": "C0392B",
    "status_change": "B8860B", "priority_change": "B8860B",
    "sms_sent": "138496", "pending": "B8860B",
}
_ACT_LBL = {
    "created": "✦ Created",  "assigned": "→ Assigned",  "reassigned": "⇄ Reassigned",
    "comment": "💬 Comment", "attachment": "📎 Attach",  "resolved": "✔ Resolved",
    "closed": "🔒 Closed",   "reopened": "↺ Reopened",   "escalated": "⚠ Escalated",
    "status_change": "⊙ Status", "priority_change": "⊙ Priority",
    "sms_sent": "✉ SMS",      "pending": "⏸ Pending",
}


def _xs(c="D4DDD0"): return Side(style="thin", color=c)
def _xb(): x = _xs(); return Border(left=x, right=x, top=x, bottom=x)
def _xf(h): return PatternFill("solid", fgColor=h)
def _xn(bold=False, color=_NAVY, size=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def _xa(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _strip(t): return re.sub(r"<[^>]+>", "", t or "")


# ─────────────────────────────────────────────────────────────────────────────
#  IMAGE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_img(url, timeout=6):
    if not _PIL_OK or not url:
        return None
    try:
        import urllib.request as _ur
        req = _ur.Request(url, headers={"User-Agent": "Mozilla/5.0 (KSTS/1.0)"})
        with _ur.urlopen(req, timeout=timeout) as r:
            data = r.read()
        img = PILImage.open(io.BytesIO(data))
        img.load()
        return img
    except Exception:
        return None


def _resize(img, max_w=160, max_h=120):
    ow, oh = img.size
    ratio = min(max_w / ow, max_h / oh, 1.0)
    nw, nh = max(1, int(ow * ratio)), max(1, int(oh * ratio))
    return img.resize((nw, nh), PILImage.LANCZOS), nw, nh


def _to_png(img):
    buf = io.BytesIO()
    img.convert("RGB").save(buf, "PNG")
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
#  ACTIVITY SERIALISER
# ─────────────────────────────────────────────────────────────────────────────

def _get_ticket_events(ticket):
    from .models import TicketCommentAttachment
    from django.db.models import Prefetch

    activities = (
        ticket.activities
        .select_related("performed_by", "comment")
        .prefetch_related(
            "assigned_to",
            Prefetch(
                "comment__attachments",
                queryset=TicketCommentAttachment.objects.all().order_by("uploaded_at"),
            ),
        )
        .order_by("created_at")
    )

    events = []
    for act in activities:
        actor  = act.performed_by.get_full_name() if act.performed_by else "System"
        atype  = act.activity_type
        anames = ", ".join(u.get_full_name() for u in act.assigned_to.all()) or "—"
        text_map = {
            "created":         f"Ticket created by {actor}",
            "assigned":        f"Assigned to {anames} by {actor}",
            "reassigned":      f"Reassigned to {anames} by {actor}",
            "status_change":   f"Status → {act.new_status or '?'} by {actor}",
            "priority_change": f"Priority → {act.new_priority or '?'} by {actor}",
            "comment":         f"Comment by {actor}",
            "attachment":      f"Attachment added by {actor}",
            "escalated":       f"ESCALATED by {actor}",
            "resolved":        f"Resolved by {actor}",
            "reopened":        f"Reopened by {actor}",
            "pending":         f"Marked Pending by {actor}",
            "closed":          f"Closed by {actor}",
            "sms_sent":        f"SMS sent to {act.sms_recipient or 'caller'}",
        }
        text = text_map.get(atype, act.description or f"Activity by {actor}")

        ev = {
            "type": atype, "text": text, "actor": actor,
            "time": act.created_at.strftime("%d %b %Y, %I:%M %p"),
            "note": "", "body_hindi": "", "attachments": [],
        }
        if act.comment:
            ev["note"]       = act.comment.body_text or act.comment.body_html or ""
            ev["body_hindi"] = act.comment.body_hindi or ""
            try:
                for att in act.comment.attachments.all():
                    ev["attachments"].append({
                        "file_name":         att.file_name,
                        "url":               att.file.url if att.file else "",
                        "is_image":          att.is_image,
                        "file_size_display": att.file_size_display,
                    })
            except Exception:
                pass

        if atype == "attachment":
            recent = ticket.attachments.all().order_by("-uploaded_at")[:10]
            seen   = {a["file_name"] for a in ev["attachments"]}
            for att in recent:
                if att.file_name not in seen:
                    ev["attachments"].append({
                        "file_name":         att.file_name,
                        "url":               att.file.url if att.file else "",
                        "is_image":          att.is_image,
                        "file_size_display": att.file_size_display,
                    })
        events.append(ev)
    return events


# ─────────────────────────────────────────────────────────────────────────────
#  SLA + OVERDUE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

_SLA_HOURS = {"critical": 24, "high": 30, "medium": 48, "low": 72}


def _sla_calc(ticket, now):
    hours    = _SLA_HOURS.get(ticket.priority, 72)
    deadline = ticket.created_at + timedelta(hours=hours)
    end_time = ticket.resolved_at if ticket.resolved_at else now
    if end_time <= deadline:
        taken = round((end_time - ticket.created_at).total_seconds() / 3600, 1)
        return f"Within SLA ({taken} hrs)", False, hours
    over = round((end_time - deadline).total_seconds() / 3600, 1)
    return f"Breached ({over} hrs over)", True, hours


def _calc_overdue_hrs(ticket, now):
    """Returns hours overdue (float). 0 if not overdue or no deadline."""
    if not ticket.expected_resolution_date:
        return 0.0
    # Consider overdue after end-of-day on the expected date
    deadline_dt = timezone.make_aware(
        timezone.datetime.combine(
            ticket.expected_resolution_date,
            timezone.datetime.min.time().replace(hour=23, minute=59, second=59),
        ),
        timezone.get_current_timezone(),
    )
    if now > deadline_dt:
        return round((now - deadline_dt).total_seconds() / 3600, 1)
    return 0.0


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

_COLS = [
    ("#",              "A",  4,  "center"),
    ("Ticket ID",      "B",  15, "center"),
    ("Caller",         "C",  20, "left"),
    ("Location",       "D",  18, "left"),
    ("Mobile",         "E",  13, "center"),
    ("Type",           "F",  20, "left"),
    ("Priority",       "G",  11, "center"),
    ("Status",         "H",  11, "center"),
    ("Escalated",      "I",   9, "center"),
    ("Overdue By",     "J",  14, "center"),
    ("Assigned To",    "K",  22, "left"),
    ("Created By",     "L",  18, "left"),
    ("Created At",     "M",  18, "center"),
    ("Expected",       "N",  14, "center"),
    ("SLA Status",     "O",  22, "center"),
    ("Desc (EN)",      "P",  32, "left"),
    ("Desc (हिंदी)", "Q",  32, "left"),
    ("Attachments",    "R",  28, "left"),
]
_NC = len(_COLS)  # 18


def _write_sheet_header(ws, tier, n_tickets, now):
    tier_label = "Reporting Manager Alert — Tier 1" if tier == 1 else "⚠ C.E. URGENT — Tier 2"
    tier_color = "B8860B" if tier == 1 else "C0392B"

    ws.merge_cells(f"A1:{get_column_letter(_NC)}1")
    c = ws["A1"]
    c.value     = f"KSTS — Escalated / Overdue Ticket Report  [{tier_label}]"
    c.font      = Font(name="Calibri", bold=True, color=_WHITE, size=13)
    c.fill      = _xf(tier_color)
    c.alignment = _xa("center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{get_column_letter(_NC)}2")
    c = ws["A2"]
    c.value = (
        f"Generated: {now.strftime('%d %b %Y, %I:%M %p')}   |   "
        f"Total tickets: {n_tickets}   |   Tier: {tier}   |   KSTS · Shwetdhara Dairy"
    )
    c.font      = Font(name="Calibri", size=8, italic=True, color=_WHITE)
    c.fill      = _xf(_G_MID)
    c.alignment = _xa("left")
    ws.row_dimensions[2].height = 14

    ws.merge_cells(f"A3:{get_column_letter(_NC)}3")
    c = ws["A3"]
    c.value = "SLA: Critical=24 hrs  |  High=30 hrs  |  Medium=48 hrs  |  Low=72 hrs"
    c.font  = Font(name="Calibri", size=8, italic=True, color=_GOLD)
    c.fill  = _xf(_CREAM); c.alignment = _xa("center")
    ws.row_dimensions[3].height = 13

    for idx, (lbl, col, w, _) in enumerate(_COLS, 1):
        c = ws.cell(row=4, column=idx)
        c.value = lbl; c.font = _xn(True, _NAVY, 9)
        c.fill  = _xf(_G_LIGHT); c.alignment = _xa("center"); c.border = _xb()
        ws.column_dimensions[col].width = w
    ws.row_dimensions[4].height = 16
    ws.freeze_panes = "A5"


def _write_ticket_row(ws, row, num, ticket, sla_label, sla_breached, overdue_hrs):
    bg = _WHITE if num % 2 == 1 else _ALT
    pk = ticket.priority.lower()
    pb, pf = _PRIO_C.get(pk, ("F1F5F9", "475569"))
    sb, sf = (_RED_BG, _RED_FG) if ticket.is_escalated else (
              (_AMB_BG, _AMB_FG) if overdue_hrs > 0 else ("DBEAFE", "1E40AF"))
    slabg, slafg = ("FFEBEE", "B71C1C") if sla_breached else ("E8F5E9", "1B5E20")
    esc           = ticket.is_escalated
    overdue_str   = f"{overdue_hrs} hrs" if overdue_hrs > 0 else "—"
    assigned      = ", ".join(u.get_full_name() for u in ticket.assigned_to.all()) or "Unassigned"

    values = [
        num, ticket.ticket_id,
        ticket.caller_display_name, ticket.caller_location,
        ticket.caller_contact_mobile or "—", ticket.ticket_type,
        ticket.priority.capitalize(), ticket.status.capitalize(),
        "⚠ Yes" if esc else "No", overdue_str,
        assigned,
        ticket.created_by.get_full_name() if ticket.created_by else "System",
        ticket.created_at.strftime("%d %b %Y, %I:%M %p") if ticket.created_at else "—",
        str(ticket.expected_resolution_date) if ticket.expected_resolution_date else "—",
        sla_label,
        ticket.description_en or "—",
        ticket.description_hi or "",
        "",
    ]
    center_cols = {1, 2, 5, 7, 8, 9, 10, 13, 14, 15}
    for idx, val in enumerate(values, 1):
        c = ws.cell(row=row, column=idx)
        c.value = val; c.border = _xb(); c.fill = _xf(bg); c.font = _xn(color=_SLATE)
        c.alignment = _xa(h="center" if idx in center_cols else "left",
                          wrap=(idx in {3, 4, 11, 12, 16, 17}))
        if idx == 7:
            c.fill = _xf(pb); c.font = _xn(True, pf)
        elif idx == 8:
            c.fill = _xf(sb); c.font = _xn(True, sf)
        elif idx == 9 and esc:
            c.fill = _xf(_RED_BG); c.font = _xn(True, _RED_FG)
        elif idx == 10 and overdue_hrs > 0:
            c.fill = _xf(_AMB_BG); c.font = _xn(True, _AMB_FG)
        elif idx == 15:
            c.fill = _xf(slabg); c.font = _xn(True, slafg)
        elif idx == 17 and val:
            c.fill = _xf(_CREAM); c.font = _xn(False, _GOLD, 8, italic=True)

    ws.row_dimensions[row].height = 32
    return row + 1


def _write_tl_header(ws, row, ticket_id, reason, is_auto):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NC)
    c = ws.cell(row=row, column=1)
    prefix    = "⏰ Auto-escalated (Overdue)" if is_auto else "⚠ Manually Escalated"
    reason_tx = f"   Reason: {reason}" if reason else ""
    c.value     = f"  ▸ Activity Timeline — {ticket_id}   [{prefix}]{reason_tx}"
    c.font      = Font(name="Calibri", bold=True, color=_WHITE, size=8)
    c.fill      = _xf("C0392B" if not is_auto else "B8860B")
    c.alignment = _xa("left")
    ws.row_dimensions[row].height = 13
    return row + 1


def _write_tl_events(ws, start_row, events):
    cur = start_row
    for ev_idx, ev in enumerate(events):
        atype = ev.get("type") or "other"
        dot   = _ACT_DOT.get(atype, "435663")
        lbl   = _ACT_LBL.get(atype, atype.replace("_", " ").title())
        actor = ev.get("actor", "System")
        text  = _strip(ev.get("text", ""))
        note  = (ev.get("note") or "").strip()
        hindi = (ev.get("body_hindi") or "").strip()
        atts  = ev.get("attachments") or []
        ttime = ev.get("time", "")
        tl_bg = "F6FAF6" if ev_idx % 2 == 0 else "EEF4EE"

        desc = text + ("\n» " + note[:400] if note else "")

        img_atts  = [a for a in atts if a.get("is_image") and a.get("url")]
        file_atts = [a for a in atts if not (a.get("is_image") and a.get("url"))]
        chip_txt  = "\n".join(
            "📎 " + (a.get("file_name") or "file")
            + ("  [" + a.get("file_size_display", "") + "]" if a.get("file_size_display") else "")
            for a in file_atts
        )

        def _tc(col, val="", h="left", wrap=False, color=_SLATE, size=8, bold=False, italic=False):
            c = ws.cell(row=cur, column=col)
            c.value = val
            c.font  = Font(name="Calibri", size=size, color=color, bold=bold, italic=italic)
            c.fill  = _xf(tl_bg); c.alignment = _xa(h=h, wrap=wrap); c.border = _xb()

        _tc(1, "  ●", color=dot, size=7)
        _tc(2, ttime, h="center", color="7F8C8D", size=7)

        c3 = ws.cell(row=cur, column=3)
        c3.value = f"{lbl}\n{actor}"
        c3.font  = Font(name="Calibri", size=7, color=dot)
        c3.fill  = _xf(tl_bg); c3.alignment = _xa("left", wrap=True); c3.border = _xb()

        ws.merge_cells(start_row=cur, start_column=4, end_row=cur, end_column=17)
        _tc(4, desc, h="left", wrap=True, color=_NAVY, size=8)
        _tc(18, chip_txt, h="left", wrap=True, color="1E55A3", size=7)

        lines = max(1, len(desc.split("\n")), len(chip_txt.split("\n")) if chip_txt else 1)
        ws.row_dimensions[cur].height = max(20, lines * 11 + 4)

        img_row = cur
        for ia in img_atts:
            pil = _fetch_img(ia.get("url", ""))
            if pil:
                resized, nw, nh = _resize(pil)
                buf = _to_png(resized)
                xi = XLImage(buf); xi.width = nw; xi.height = nh
                ws.add_image(xi, f"{get_column_letter(18)}{img_row}")
                needed = nh * 0.75 + 4
                if (ws.row_dimensions[img_row].height or 0) < needed:
                    ws.row_dimensions[img_row].height = needed
                img_row += 1
            else:
                name = ia.get("file_name", "image")
                prev = ws.cell(row=img_row, column=18).value or ""
                ws.cell(row=img_row, column=18).value = (prev + "\n" if prev else "") + f"🖼 {name}"

        cur = max(cur + 1, img_row)

        if hindi:
            ws.merge_cells(start_row=cur, start_column=4, end_row=cur, end_column=18)
            hc = ws.cell(row=cur, column=4)
            hc.value = "हिंदी: " + hindi[:400]
            hc.font  = Font(name="Calibri", size=7, color=_GOLD, italic=True)
            hc.fill  = _xf(_CREAM); hc.alignment = _xa("left", wrap=True); hc.border = _xb()
            for col in [1, 2, 3]:
                cx = ws.cell(row=cur, column=col)
                cx.fill = _xf(_CREAM); cx.border = _xb()
            ws.row_dimensions[cur].height = max(13, len(hindi) // 8 + 12)
            cur += 1

    for col in range(1, _NC + 1):
        c = ws.cell(row=cur, column=col)
        c.fill = _xf(_G_LIGHT); c.border = _xb(); c.value = ""
    ws.row_dimensions[cur].height = 3
    return cur + 1


def _build_summary_sheet(wb, rows, tier):
    tier_color = "B8860B" if tier == 1 else "C0392B"
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    ws.merge_cells("A1:B1")
    c = ws["A1"]; c.value = f"Escalation Report Summary — Tier {tier}"
    c.font = Font(name="Calibri", bold=True, color=_WHITE, size=12)
    c.fill = _xf(tier_color); c.alignment = _xa("center")
    ws.row_dimensions[1].height = 22

    prio_c = Counter(r["priority"].lower() for r in rows)
    type_c = Counter(r["ticket_type"]      for r in rows)
    mode_c = Counter(
        "Auto-escalated (Overdue)" if r.get("auto_escalated") else "Manually Escalated"
        for r in rows
    )

    row = 2
    for sec, counter in [
        ("By Priority",         prio_c),
        ("By Ticket Type",      type_c),
        ("By Escalation Mode",  mode_c),
    ]:
        ws.merge_cells(f"A{row}:B{row}")
        h = ws.cell(row=row, column=1); h.value = sec
        h.font = Font(name="Calibri", bold=True, color=_NAVY, size=10)
        h.fill = _xf(_G_LIGHT); h.alignment = _xa("left")
        row += 1
        for key, cnt in sorted(counter.items(), key=lambda x: -x[1]):
            bg = _WHITE if row % 2 == 0 else _ALT
            la = ws.cell(row=row, column=1); la.value = key.capitalize()
            lb = ws.cell(row=row, column=2); lb.value = cnt
            for c in (la, lb):
                c.fill = _xf(bg); c.font = Font(name="Calibri", size=9, color=_SLATE); c.border = _xb()
            la.alignment = _xa("left"); lb.alignment = _xa("center")
            row += 1
        row += 1


def build_escalation_excel(ticket_orm_list, tier, ticket_meta_list=None):
    """
    Returns in-memory .xlsx bytes for email attachment.
    ticket_meta_list: list of dicts with keys auto_escalated, overdue_hrs.
    """
    now = timezone.now()
    if ticket_meta_list is None:
        ticket_meta_list = [{"auto_escalated": False, "overdue_hrs": 0.0}] * len(ticket_orm_list)

    wb = Workbook()
    ws = wb.active
    ws.title = "Escalated Tickets"
    ws.sheet_view.showGridLines = False

    _write_sheet_header(ws, tier, len(ticket_orm_list), now)

    cur = 5
    summary_rows = []
    for num, (ticket, meta) in enumerate(zip(ticket_orm_list, ticket_meta_list), 1):
        sla_label, sla_breached, _ = _sla_calc(ticket, now)
        overdue_hrs = meta.get("overdue_hrs", 0.0)
        cur = _write_ticket_row(ws, cur, num, ticket, sla_label, sla_breached, overdue_hrs)
        cur = _write_tl_header(ws, cur, ticket.ticket_id,
                                (ticket.escalation_reason or "").strip(),
                                meta.get("auto_escalated", False))
        cur = _write_tl_events(ws, cur, _get_ticket_events(ticket))
        summary_rows.append({
            "priority":       ticket.priority,
            "ticket_type":    ticket.ticket_type,
            "auto_escalated": meta.get("auto_escalated", False),
        })

    _build_summary_sheet(wb, summary_rows, tier)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ─────────────────────────────────────────────────────────────────────────────
#  EMAIL HTML
# ─────────────────────────────────────────────────────────────────────────────

def _ticket_url(tid):
    return f"{SITE_URL.rstrip('/')}/home/?ticket={tid}"


def _build_tier1_html(tickets_with_meta, manager_name):
    rows = ""
    for ticket, meta in tickets_with_meta:
        is_auto   = meta.get("auto_escalated", False)
        over_h    = meta.get("overdue_hrs", 0.0)
        badge_txt = "⏰ Auto (Overdue)" if is_auto else "⚠ Manual"
        badge_col = "#854D0E" if is_auto else "#991B1B"
        badge_bg  = "#FEF9C3" if is_auto else "#FEE2E2"
        over_td   = (f'<span style="color:#C0392B;font-weight:700">{over_h} hrs overdue</span>'
                     if over_h > 0 else "—")
        rows += f"""
        <tr>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;font-size:13px">
            <a href="{_ticket_url(ticket.ticket_id)}" style="color:#2F6B3F;font-weight:700">{ticket.ticket_id}</a>
          </td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;font-size:13px">{ticket.caller_display_name}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;font-size:13px">{ticket.ticket_type}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:center">
            <span style="padding:3px 10px;border-radius:12px;background:{badge_bg};color:{badge_col};font-weight:700;font-size:11px">{badge_txt}</span>
          </td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:center">
            <span style="padding:3px 10px;border-radius:12px;background:#FEE2E2;color:#991B1B;font-weight:700;font-size:11px">{ticket.priority.upper()}</span>
          </td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;font-size:12px">{over_td}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;font-size:12px;color:#666">{(ticket.escalation_reason or '—')[:80]}</td>
        </tr>"""

    auto_count   = sum(1 for _, m in tickets_with_meta if m.get("auto_escalated"))
    manual_count = len(tickets_with_meta) - auto_count
    badges = ""
    if auto_count:
        badges += f'<span style="background:#FEF9C3;color:#854D0E;font-weight:700;padding:2px 8px;border-radius:6px">⏰ {auto_count} auto-escalated (overdue)</span> '
    if manual_count:
        badges += f'<span style="background:#FEE2E2;color:#991B1B;font-weight:700;padding:2px 8px;border-radius:6px;margin-left:4px">⚠ {manual_count} manually escalated</span>'

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f5f7f3;font-family:Calibri,Arial,sans-serif">
<div style="max-width:780px;margin:24px auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.10)">
  <div style="background:linear-gradient(135deg,#b8860b,#d4a017);padding:28px 32px">
    <div style="font-size:22px;font-weight:700;color:white">⚠ Escalated / Overdue Ticket Alert</div>
    <div style="font-size:13px;color:rgba(255,255,255,.8);margin-top:4px">KSTS · Shwetdhara Dairy</div>
  </div>
  <div style="padding:24px 32px 0">
    <p style="font-size:15px;color:#2C3547;margin:0">Dear <strong>{manager_name}</strong>,</p>
    <p style="font-size:14px;color:#435663;margin-top:10px;line-height:1.6">
      The following <strong>{len(tickets_with_meta)} ticket{'s' if len(tickets_with_meta)>1 else ''}</strong>
      require <strong style="color:#C0392B">immediate attention</strong>. {badges}
    </p>
    <div style="background:#FFF8DC;border-left:4px solid #b8860b;padding:11px 14px;border-radius:0 8px 8px 0;margin-top:14px;font-size:13px;color:#856404">
      ⏰ If unresolved, the C.E. will be automatically notified after <strong>{TIER2_DELAY_HOURS} hours</strong>.
    </div>
  </div>
  <div style="padding:20px 32px">
    <table style="width:100%;border-collapse:collapse;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">
      <thead><tr style="background:#2C3547">
        <th style="padding:10px 12px;text-align:left;color:#C5D9C0;font-size:12px">Ticket ID</th>
        <th style="padding:10px 12px;text-align:left;color:#C5D9C0;font-size:12px">Caller</th>
        <th style="padding:10px 12px;text-align:left;color:#C5D9C0;font-size:12px">Type</th>
        <th style="padding:10px 12px;text-align:center;color:#C5D9C0;font-size:12px">Mode</th>
        <th style="padding:10px 12px;text-align:center;color:#C5D9C0;font-size:12px">Priority</th>
        <th style="padding:10px 12px;text-align:left;color:#C5D9C0;font-size:12px">Overdue</th>
        <th style="padding:10px 12px;text-align:left;color:#C5D9C0;font-size:12px">Reason</th>
      </tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
  <div style="padding:0 32px 28px;text-align:center">
    <a href="{SITE_URL.rstrip('/')}/home/" style="display:inline-block;padding:12px 32px;background:linear-gradient(135deg,#2F6B3F,#3a7d4f);color:white;font-weight:700;font-size:14px;border-radius:8px;text-decoration:none">
      Open KSTS Dashboard →
    </a>
  </div>
  <div style="padding:16px 32px;background:#f8faf6;border-top:1px solid #e2e8f0">
    <p style="font-size:11px;color:#9CA3AF;margin:0">Automated notification · KSTS · Shwetdhara Dairy · Full timeline in attached Excel.</p>
  </div>
</div></body></html>"""


def _build_tier2_html(tickets_with_meta, ce_name=""):
    rows = ""
    now = timezone.now()
    for ticket, meta in tickets_with_meta:
        over_h = meta.get("overdue_hrs", 0.0)
        elapsed_t1 = "—"
        try:
            notif = ticket.escalation_notification
            if notif.tier1_sent_at:
                hrs = round((now - notif.tier1_sent_at).total_seconds() / 3600, 1)
                elapsed_t1 = f"{hrs} hrs ago"
        except Exception:
            pass
        assigned = ", ".join(u.get_full_name() for u in ticket.assigned_to.all()) or "Unassigned"
        over_td  = (f'<strong style="color:#C0392B">{over_h} hrs</strong>' if over_h > 0 else "—")
        rows += f"""
        <tr>
          <td style="padding:8px 12px;border-bottom:1px solid #fecaca;font-size:13px">
            <a href="{_ticket_url(ticket.ticket_id)}" style="color:#991B1B;font-weight:700">{ticket.ticket_id}</a>
          </td>
          <td style="padding:8px 12px;border-bottom:1px solid #fecaca;font-size:13px">{ticket.caller_display_name}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #fecaca;font-size:13px">{ticket.ticket_type}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #fecaca;text-align:center">
            <span style="padding:3px 9px;border-radius:12px;background:#FEE2E2;color:#991B1B;font-weight:700;font-size:11px">{ticket.priority.upper()}</span>
          </td>
          <td style="padding:8px 12px;border-bottom:1px solid #fecaca;font-size:12px">{over_td}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #fecaca;font-size:12px;color:#C0392B">{elapsed_t1}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #fecaca;font-size:12px">{assigned}</td>
        </tr>"""

    greeting = f"Dear <strong>{ce_name}</strong>," if ce_name else "Dear Sir/Ma'am,"
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f5f7f3;font-family:Calibri,Arial,sans-serif">
<div style="max-width:780px;margin:24px auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.12)">
  <div style="background:linear-gradient(135deg,#7f1d1d,#C0392B);padding:28px 32px">
    <div style="font-size:24px;font-weight:700;color:white">🚨 URGENT — Unresolved Escalated Tickets</div>
    <div style="font-size:13px;color:rgba(255,255,255,.8);margin-top:4px">Chief Executive Notification · KSTS · Shwetdhara Dairy</div>
  </div>
  <div style="padding:24px 32px 0">
    <p style="font-size:15px;color:#2C3547;margin:0">{greeting}</p>
    <p style="font-size:14px;color:#435663;margin-top:10px;line-height:1.6">
      This is an <strong style="color:#C0392B">automated Tier-2 escalation</strong>.
      The following <strong>{len(tickets_with_meta)} ticket{'s' if len(tickets_with_meta)>1 else ''}</strong>
      remained <strong>unresolved for over {TIER2_DELAY_HOURS} hours</strong> after manager notification.
    </p>
    <div style="background:#FEE2E2;border-left:4px solid #C0392B;padding:12px 16px;border-radius:0 8px 8px 0;margin-top:14px;font-size:13px;color:#7f1d1d;font-weight:600">
      ⚠ Immediate executive action or delegation is required.
    </div>
  </div>
  <div style="padding:20px 32px">
    <table style="width:100%;border-collapse:collapse;border:1px solid #fecaca;border-radius:8px;overflow:hidden">
      <thead><tr style="background:#7f1d1d">
        <th style="padding:10px 12px;text-align:left;color:#fecaca;font-size:12px">Ticket ID</th>
        <th style="padding:10px 12px;text-align:left;color:#fecaca;font-size:12px">Caller</th>
        <th style="padding:10px 12px;text-align:left;color:#fecaca;font-size:12px">Type</th>
        <th style="padding:10px 12px;text-align:center;color:#fecaca;font-size:12px">Priority</th>
        <th style="padding:10px 12px;text-align:left;color:#fecaca;font-size:12px">Overdue</th>
        <th style="padding:10px 12px;text-align:left;color:#fecaca;font-size:12px">Since Mgr Alert</th>
        <th style="padding:10px 12px;text-align:left;color:#fecaca;font-size:12px">Assigned To</th>
      </tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
  <div style="padding:0 32px 28px;text-align:center">
    <a href="{SITE_URL.rstrip('/')}/home/" style="display:inline-block;padding:12px 32px;background:#C0392B;color:white;font-weight:700;font-size:14px;border-radius:8px;text-decoration:none">
      Open Dashboard — Take Action →
    </a>
  </div>
  <div style="padding:16px 32px;background:#fff5f5;border-top:1px solid #fecaca">
    <p style="font-size:11px;color:#9CA3AF;margin:0">Automated Tier-2 escalation · KSTS · Shwetdhara Dairy · Full history in attached Excel.</p>
  </div>
</div></body></html>"""


# ─────────────────────────────────────────────────────────────────────────────
#  DISPATCH: TIER-1
# ─────────────────────────────────────────────────────────────────────────────

# In escalation_engine.py - Replace the dispatch_tier1_for_ticket function

def dispatch_tier1_for_ticket(ticket_id, is_auto_escalated=False, overdue_hrs=0.0):
    from .models import Ticket, EscalationNotification

    try:
        ticket = (
            Ticket.objects
            .select_related("created_by", "escalated_to")
            .prefetch_related(
                "assigned_to__manager",
                "activities__performed_by",
                "activities__comment__attachments",
                "activities__assigned_to",
                "attachments",
            )
            .get(ticket_id=ticket_id)
        )
    except Ticket.DoesNotExist:
        logger.error("Tier-1: ticket %s not found", ticket_id)
        return

    notif, _ = EscalationNotification.objects.get_or_create(ticket=ticket)

    if notif.tier1_sent:
        logger.info("Tier-1 already sent for %s — skipping", ticket_id)
        return

    # === NEW: Collect ALL recipients (users + their HODs) ===
    recipients = {}  # email -> {"name": str, "type": str}
    
    # 1. Add all assigned users
    for assignee in ticket.assigned_to.all():
        if assignee.email:
            recipients[assignee.email] = {
                "name": assignee.get_full_name() or assignee.email,
                "type": "assignee"
            }
    
    # 2. Add all managers/HODs of assigned users
    for assignee in ticket.assigned_to.all():
        mgr = assignee.manager
        if mgr and mgr.email and mgr.email not in recipients:
            recipients[mgr.email] = {
                "name": mgr.get_full_name() or mgr.email,
                "type": "hod"
            }
    
    # 3. Add escalated_to user if different from assignees
    if ticket.escalated_to and ticket.escalated_to.email:
        if ticket.escalated_to.email not in recipients:
            recipients[ticket.escalated_to.email] = {
                "name": ticket.escalated_to.get_full_name() or ticket.escalated_to.email,
                "type": "escalated_to"
            }
    
    # 4. Add escalated_to's manager/HOD
    if ticket.escalated_to and ticket.escalated_to.manager:
        mgr = ticket.escalated_to.manager
        if mgr and mgr.email and mgr.email not in recipients:
            recipients[mgr.email] = {
                "name": mgr.get_full_name() or mgr.email,
                "type": "hod"
            }

    if not recipients:
        logger.warning("No recipients found for %s", ticket_id)
        # Fallback to CE
        dispatch_tier2_for_ticket(ticket_id, force=True,
                                   is_auto_escalated=is_auto_escalated,
                                   overdue_hrs=overdue_hrs)
        return

    meta = {"auto_escalated": is_auto_escalated, "overdue_hrs": overdue_hrs}
    excel_bytes = build_escalation_excel([ticket], tier=1, ticket_meta_list=[meta])
    filename = f"Escalation_T1_{ticket_id}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    sent_emails = []
    
    # Send to each recipient
    for email, recipient_info in recipients.items():
        recipient_type = recipient_info["type"]
        recipient_name = recipient_info["name"]
        
        # Customize email based on recipient type
        if recipient_type == "assignee":
            subject = f"⚠ [KSTS] Ticket Escalated — Action Required from You — {ticket_id}"
            greeting = f"Dear {recipient_name},"
            intro_text = (
                f"A ticket you are assigned to has been <strong>escalated</strong>. "
                f"Please take immediate action. The ticket has been overdue by "
                f"<strong>{overdue_hrs:.1f} hours</strong>."
            )
            cta_text = "View & Update Ticket"
        elif recipient_type == "hod":
            subject = f"⚠ [KSTS] Escalated Ticket Alert — Action Required from HOD — {ticket_id}"
            greeting = f"Dear {recipient_name} (HOD),"
            intro_text = (
                f"A ticket under your team has been <strong>escalated</strong>. "
                f"Please ensure your team member takes immediate action. "
                f"The ticket has been overdue by <strong>{overdue_hrs:.1f} hours</strong>."
            )
            cta_text = "Review Team Ticket"
        else:
            subject = f"⚠ [KSTS] Escalated Ticket Alert — Immediate Attention Required — {ticket_id}"
            greeting = f"Dear {recipient_name},"
            intro_text = (
                f"A support ticket has been <strong>escalated</strong> and requires attention. "
                f"The ticket has been overdue by <strong>{overdue_hrs:.1f} hours</strong>."
            )
            cta_text = "Review Ticket"
        
        # Build custom HTML for this recipient
        html = _build_tier1_html_for_recipient(
            ticket=ticket,
            meta=meta,
            recipient_name=recipient_name,
            recipient_type=recipient_type,
            greeting=greeting,
            intro_text=intro_text,
            cta_text=cta_text,
            overdue_hrs=overdue_hrs
        )
        
        try:
            msg = EmailMessage(
                subject=subject,
                body=html,
                from_email=f"{SENDER_NAME} <{settings.DEFAULT_FROM_EMAIL}>",
                to=[email],
            )
            msg.content_subtype = "html"
            msg.attach(filename, excel_bytes,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            msg.send(fail_silently=False)
            sent_emails.append(email)
            logger.info("Tier-1 sent: %s → %s (type=%s)", ticket_id, email, recipient_type)
        except Exception as exc:
            logger.error("Tier-1 FAILED: %s → %s: %s", ticket_id, email, exc)

    if sent_emails:
        notif.tier1_sent_at = timezone.now()
        notif.tier1_recipient = ", ".join(sent_emails)
        notif.save(update_fields=["tier1_sent_at", "tier1_recipient", "updated_at"])
        logger.info("Tier-1 recorded for %s → %s", ticket_id, notif.tier1_recipient)


def _build_tier1_html_for_recipient(ticket, meta, recipient_name, recipient_type, 
                                     greeting, intro_text, cta_text, overdue_hrs):
    """Build HTML email customized for recipient type."""
    is_auto = meta.get("auto_escalated", False)
    badge_txt = "⏰ Auto (Overdue)" if is_auto else "⚠ Manual"
    badge_col = "#854D0E" if is_auto else "#991B1B"
    badge_bg = "#FEF9C3" if is_auto else "#FEE2E2"
    
    assigned_users = ", ".join(u.get_full_name() for u in ticket.assigned_to.all()) or "Unassigned"
    
    # Determine action message based on recipient type
    if recipient_type == "assignee":
        action_msg = "Please update the ticket status or add comments with your action plan."
    elif recipient_type == "hod":
        action_msg = "Please review this escalated ticket and guide your team member for resolution."
    else:
        action_msg = "Please review this escalated ticket and take appropriate action."
    
    return f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(135deg, #b8860b, #d4a017); padding: 20px; color: white; }}
        .badge {{ display: inline-block; background: {badge_bg}; color: {badge_col}; 
                  padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: bold; }}
        .content {{ padding: 20px; background: #f9f9f9; }}
        .ticket-details {{ background: white; padding: 15px; border-radius: 8px; margin: 15px 0; }}
        .detail-row {{ padding: 8px 0; border-bottom: 1px solid #eee; }}
        .label {{ font-weight: bold; width: 120px; display: inline-block; }}
        .button {{ display: inline-block; background: #b8860b; color: white; padding: 12px 24px; 
                  text-decoration: none; border-radius: 6px; margin-top: 20px; }}
        .footer {{ font-size: 11px; color: #999; text-align: center; margin-top: 20px; 
                  padding-top: 20px; border-top: 1px solid #eee; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <span class="badge">{badge_txt}</span>
            <h2>⚠ Ticket Escalation Alert</h2>
            <p>KSTS · Shwetdhara Dairy</p>
        </div>
        <div class="content">
            <p>{greeting}</p>
            <p>{intro_text}</p>
            
            <div class="ticket-details">
                <div class="detail-row">
                    <span class="label">Ticket ID:</span>
                    <span><strong>{ticket.ticket_id}</strong></span>
                </div>
                <div class="detail-row">
                    <span class="label">Priority:</span>
                    <span><strong style="color: #d32f2f;">{ticket.priority.upper()}</strong></span>
                </div>
                <div class="detail-row">
                    <span class="label">Caller:</span>
                    <span>{ticket.caller_display_name}</span>
                </div>
                <div class="detail-row">
                    <span class="label">Assigned To:</span>
                    <span>{assigned_users}</span>
                </div>
                <div class="detail-row">
                    <span class="label">Overdue By:</span>
                    <span><strong style="color: #d32f2f;">{overdue_hrs:.1f} hours</strong></span>
                </div>
                <div class="detail-row">
                    <span class="label">Reason:</span>
                    <span>{ticket.escalation_reason or 'Not specified'}</span>
                </div>
            </div>
            
            <p><strong>What you need to do:</strong><br>
            {action_msg}</p>
            
            <div style="text-align: center;">
                <a href="{SITE_URL}/home/?ticket={ticket.ticket_id}" class="button">{cta_text} →</a>
            </div>
        </div>
        <div class="footer">
            <p>This is an automated notification from KSTS. Please do not reply to this email.</p>
            <p>Generated: {timezone.now().strftime('%d %b %Y, %I:%M %p')}</p>
        </div>
    </div>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
#  DISPATCH: TIER-2
# ─────────────────────────────────────────────────────────────────────────────

# In escalation_engine.py - Update dispatch_tier2_for_ticket

def dispatch_tier2_for_ticket(ticket_id, force=False,
                               is_auto_escalated=False, overdue_hrs=0.0):
    from .models import Ticket, EscalationNotification, CustomUser

    if not CE_EMAIL:
        logger.error("ESCALATION_CE_EMAIL not set — Tier-2 aborted for %s", ticket_id)
        return

    try:
        ticket = (
            Ticket.objects
            .select_related("created_by", "escalated_to")
            .prefetch_related(
                "assigned_to__manager",
                "activities__performed_by",
                "activities__comment__attachments",
                "activities__assigned_to",
                "attachments",
                "escalation_notification",
            )
            .get(ticket_id=ticket_id)
        )
    except Ticket.DoesNotExist:
        logger.error("Tier-2: ticket %s not found", ticket_id)
        return

    if ticket.status in ("resolved", "closed"):
        logger.info("Ticket %s resolved — Tier-2 skipped", ticket_id)
        return

    notif, _ = EscalationNotification.objects.get_or_create(ticket=ticket)

    if notif.tier2_sent:
        logger.info("Tier-2 already sent for %s", ticket_id)
        return

    if not force:
        if not notif.tier1_sent_at:
            logger.info("Tier-2 skipped for %s — no T1 record", ticket_id)
            return
        elapsed = (timezone.now() - notif.tier1_sent_at).total_seconds() / 3600
        if elapsed < TIER2_DELAY_HOURS:
            logger.info("Tier-2 skipped for %s — %.1f hrs since T1 (need %d)",
                        ticket_id, elapsed, TIER2_DELAY_HOURS)
            return

    meta = {"auto_escalated": is_auto_escalated, "overdue_hrs": overdue_hrs}
    excel_bytes = build_escalation_excel([ticket], tier=2, ticket_meta_list=[meta])
    filename = f"Escalation_T2_CE_{ticket_id}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    ce_user = CustomUser.objects.filter(email__iexact=CE_EMAIL).first()
    ce_name = ce_user.get_full_name() if ce_user else ""

    # === NEW: Collect HODs to CC ===
    cc_emails = set()
    
    # Add all HODs from assigned users
    for assignee in ticket.assigned_to.all():
        mgr = assignee.manager
        if mgr and mgr.email:
            cc_emails.add(mgr.email)
    
    # Add escalated_to's HOD
    if ticket.escalated_to and ticket.escalated_to.manager:
        mgr = ticket.escalated_to.manager
        if mgr and mgr.email:
            cc_emails.add(mgr.email)
    
    # Add previous Tier-1 recipients (assignees + HODs)
    if notif.tier1_recipient:
        for email in notif.tier1_recipient.split(","):
            email = email.strip()
            if email:
                cc_emails.add(email)
    
    # Remove CE from CC if present
    cc_emails.discard(CE_EMAIL)
    
    cc_list = list(cc_emails)[:10]  # Limit to 10 CC recipients

    html = _build_tier2_html([(ticket, meta)], ce_name)
    try:
        msg = EmailMessage(
            subject=f"🚨 [KSTS] URGENT: Escalated Ticket Unresolved — C.E. Notification — {ticket_id}",
            body=html,
            from_email=f"{SENDER_NAME} <{settings.DEFAULT_FROM_EMAIL}>",
            to=[CE_EMAIL],
            cc=cc_list,  # HODs in CC
        )
        msg.content_subtype = "html"
        msg.attach(filename, excel_bytes,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        msg.send(fail_silently=False)

        notif.tier2_sent_at = timezone.now()
        notif.tier2_recipient = CE_EMAIL
        notif.save(update_fields=["tier2_sent_at", "tier2_recipient", "updated_at"])
        logger.info("Tier-2 CE sent: %s → %s (CC: %s)", ticket_id, CE_EMAIL, cc_list)
    except Exception as exc:
        logger.error("Tier-2 FAILED for %s: %s", ticket_id, exc)
        raise

# ─────────────────────────────────────────────────────────────────────────────
#  SWEEP A — AUTO-ESCALATE OVERDUE TICKETS  (runs every 15 min)
# ─────────────────────────────────────────────────────────────────────────────

def run_overdue_auto_escalate_sweep():
    """
    Finds all non-escalated tickets past their expected_resolution_date,
    marks them escalated, logs an activity, fires Tier-1 immediately,
    and schedules Tier-2 for TIER2_DELAY_HOURS later.
    """
    from .models import Ticket, TicketActivity, EscalationNotification
    from .tasks import send_tier2_escalation

    now   = timezone.now()
    today = now.date()

    candidates = (
        Ticket.objects
        .filter(
            status__in=_ACTIVE_STATUSES,
            is_escalated=False,
            expected_resolution_date__lt=today,
        )
        .select_related("created_by")
        .prefetch_related("assigned_to__manager", "attachments")
    )

    count = 0
    for ticket in candidates:
        overdue_h = _calc_overdue_hrs(ticket, now)
        if overdue_h <= 0:
            continue

        logger.info("Auto-escalating %s (%.1f hrs overdue, priority=%s)",
                    ticket.ticket_id, overdue_h, ticket.priority)

        old_status             = ticket.status
        ticket.is_escalated    = True
        ticket.status          = Ticket.Status.ESCALATED
        ticket.escalated_at    = now
        ticket.escalation_reason = f"Auto-escalated: overdue by {overdue_h:.1f} hrs"
        ticket.save(update_fields=[
            "is_escalated", "status", "escalated_at",
            "escalation_reason", "updated_at",
        ])

        TicketActivity.objects.create(
            ticket        = ticket,
            activity_type = TicketActivity.ActivityType.ESCALATED,
            performed_by  = None,
            old_status    = old_status,
            new_status    = Ticket.Status.ESCALATED,
            description   = f"Auto-escalated by KSTS system: overdue by {overdue_h:.1f} hrs.",
        )

        EscalationNotification.objects.get_or_create(ticket=ticket)

        try:
            dispatch_tier1_for_ticket(
                ticket.ticket_id,
                is_auto_escalated=True,
                overdue_hrs=overdue_h,
            )
        except Exception as exc:
            logger.error("T1 failed in auto-escalate sweep for %s: %s", ticket.ticket_id, exc)

        try:
            send_tier2_escalation.apply_async(
                kwargs={
                    "ticket_id":         ticket.ticket_id,
                    "is_auto_escalated": True,
                    "overdue_hrs":       overdue_h,
                },
                countdown=TIER2_DELAY_HOURS * 3600,
            )
        except Exception as exc:
            logger.error("Could not schedule T2 for %s: %s", ticket.ticket_id, exc)

        count += 1

    logger.info("Overdue sweep done — %d tickets auto-escalated", count)
    return count


# ─────────────────────────────────────────────────────────────────────────────
#  SWEEP B — TIER-2 SAFETY NET  (runs every 30 min)
# ─────────────────────────────────────────────────────────────────────────────

def run_tier2_sweep():
    from .models import EscalationNotification

    cutoff = timezone.now() - timedelta(hours=TIER2_DELAY_HOURS)
    candidates = (
        EscalationNotification.objects
        .filter(tier1_sent_at__lte=cutoff, tier2_sent_at__isnull=True)
        .select_related("ticket")
    )

    triggered = 0
    for notif in candidates:
        t = notif.ticket
        if t.status in ("resolved", "closed") or not t.is_escalated:
            continue
        logger.info("Tier-2 sweep: triggering for %s", t.ticket_id)
        try:
            dispatch_tier2_for_ticket(t.ticket_id)
            triggered += 1
        except Exception as exc:
            logger.error("T2 sweep error for %s: %s", t.ticket_id, exc)

    logger.info("Tier-2 sweep done — %d tickets processed", triggered)
    return triggered