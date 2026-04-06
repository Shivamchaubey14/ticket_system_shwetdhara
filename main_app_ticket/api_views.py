# ksts/api_views.py
import datetime
import io
import json
import logging
import mimetypes
import re
import urllib.request as _urlreq
from collections import Counter
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.db.models import Q, Prefetch
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

from .models import (
    CustomUser, Farmer, MPP, Ticket, TicketActivity,
    TicketAttachment, TicketComment, TicketCommentAttachment, Transporter,
)
from .ticket_mailer import (
    ticket_assigned_notification,
    ticket_escalation_notification,
    ticket_manager_area_alert,
    ticket_reassign_notification,
    ticket_status_notification,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
#  RESPONSE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _json_error(msg, status=400):
    return JsonResponse({"ok": False, "error": msg}, status=status)


def _json_ok(data=None, **kwargs):
    payload = {"ok": True}
    if data is not None:
        payload.update(data)
    payload.update(kwargs)
    return JsonResponse(payload)


def _q_icontains(field, term):
    return Q(**{f"{field}__icontains": term})


# ─────────────────────────────────────────────────────────────────────────────
#  MISC HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _safe_accepted_by(farmer):
    try:
        user = farmer.accepted_by
        if user is not None:
            return user.get_full_name()
    except Exception:
        pass
    return ""


def _file_type_from_name(filename):
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    if ext in ("png", "jpg", "jpeg", "gif", "webp", "svg"):
        return "image"
    if ext == "pdf":
        return "pdf"
    if ext in ("doc", "docx"):
        return "word"
    if ext in ("xls", "xlsx", "csv"):
        return "excel"
    if ext in ("ppt", "pptx"):
        return "ppt"
    if ext in ("zip", "rar", "7z"):
        return "zip"
    if ext in ("mp4", "mov", "avi", "mkv"):
        return "video"
    return "other"


def _fmt_size(b):
    if b < 1024:
        return f"{b} B"
    if b < 1_048_576:
        return f"{b/1024:.1f} KB"
    return f"{b/1_048_576:.1f} MB"


def _get_activities_qs(ticket):
    return (
        ticket.activities
        .select_related("performed_by", "comment")
        .prefetch_related(
            "assigned_to",
            Prefetch(
                "comment__attachments",
                queryset=TicketCommentAttachment.objects.all().order_by("uploaded_at"),
            ),
        )
        .order_by("created_at")
    )


def _serialize_activities(activities_qs):
    events = []
    for act in activities_qs:
        actor      = act.performed_by.get_full_name() if act.performed_by else "System"
        actor_code = act.performed_by.employee_code   if act.performed_by else ""
        atype      = act.activity_type

        if atype == "created":
            text = f"Ticket created by <strong>{actor}</strong>"
        elif atype == "assigned":
            names = ", ".join(u.get_full_name() for u in act.assigned_to.all()) or "—"
            text = f"Assigned to <strong>{names}</strong> by {actor}"
        elif atype == "reassigned":
            names = ", ".join(u.get_full_name() for u in act.assigned_to.all()) or "—"
            text = f"Reassigned to <strong>{names}</strong> by {actor}"
        elif atype == "status_change":
            text = (f"Status changed to <strong>{act.new_status}</strong> by {actor}"
                    if act.new_status else act.description or f"Ticket updated by {actor}")
        elif atype == "priority_change":
            text = f"Priority changed to <strong>{act.new_priority}</strong> by {actor}"
        elif atype == "comment":
            text = f"Comment added by <strong>{actor}</strong>"
        elif atype == "attachment":
            att_names = []
            if act.ticket and hasattr(act.ticket, "attachments"):
                recent = act.ticket.attachments.filter(
                    uploaded_at__gte=act.created_at - timezone.timedelta(seconds=5),
                    uploaded_at__lte=act.created_at + timezone.timedelta(seconds=5),
                )
                att_names = [a.file_name for a in recent]
            if att_names:
                text = f"Attachment{'s' if len(att_names)>1 else ''} added: {', '.join(att_names)} by <strong>{actor}</strong>"
            else:
                text = f"Attachment added by <strong>{actor}</strong>"
        elif atype == "escalated":
            text = f"Ticket <strong>Escalated</strong> by {actor}"
        elif atype == "resolved":
            text = f"Resolved by <strong>{actor}</strong>"
        elif atype == "reopened":
            text = f"Reopened by <strong>{actor}</strong>"
        elif atype == "pending":
            text = f"Ticket marked <strong>Pending</strong> by {actor}"
        elif atype == "closed":
            text = f"Closed by <strong>{actor}</strong>"
        elif atype == "sms_sent":
            text = f"SMS sent to <strong>{act.sms_recipient or 'caller'}</strong>"
        else:
            text = act.description or f"Activity by {actor}"

        ev = {
            "type": atype, "activity_type": atype,
            "text": text, "description": text,
            "time": act.created_at.strftime("%d %b %Y, %I:%M %p"),
            "created_at": act.created_at.strftime("%d %b %Y, %I:%M %p"),
            "actor": actor, "actor_code": actor_code,
            "note": "", "body_hindi": "", "hindi_fallback": False,
            "attachments": [],
        }

        if act.comment:
            ev["note"]           = act.comment.body_text or act.comment.body_html or ""
            ev["body_hindi"]     = act.comment.body_hindi or ""
            ev["hindi_fallback"] = act.comment.hindi_fallback
            try:
                for att in act.comment.attachments.all():
                    ev["attachments"].append({
                        "file_name": att.file_name, "name": att.file_name,
                        "url": att.file.url if att.file else "",
                        "is_image": att.is_image, "size": att.file_size_bytes,
                        "file_size_display": att.file_size_display,
                        "file_type": att.file_type, "mime_type": att.mime_type or "",
                    })
            except Exception:
                pass

        if atype == "attachment" and act.ticket:
            recent = act.ticket.attachments.all().order_by("-uploaded_at")[:10]
            existing = {a.get("file_name", "") for a in ev["attachments"]}
            for att in recent:
                if att.file_name not in existing:
                    ev["attachments"].append({
                        "file_name": att.file_name, "name": att.file_name,
                        "url": att.file.url if att.file else "",
                        "is_image": att.is_image, "size": att.file_size_bytes,
                        "file_size_display": att.file_size_display,
                        "file_type": att.file_type, "mime_type": att.mime_type or "",
                    })

        events.append(ev)
    return events


def _get_ticket_or_error(ticket_id):
    try:
        return Ticket.objects.get(ticket_id=ticket_id), None
    except Ticket.DoesNotExist:
        return None, _json_error("Ticket not found.", status=404)


# ─────────────────────────────────────────────────────────────────────────────
#  FARMER SEARCH
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def farmer_search(request):
    raw   = request.GET.get("q", "").strip()
    limit = min(int(request.GET.get("limit", 30)), 200)
    if not raw:
        return _json_ok(results=[])
    tokens = [t for t in raw.split() if len(t) >= 2]
    if not tokens:
        return _json_ok(results=[])

    qs = Farmer.objects.select_related(
        "village", "tehsil", "district", "state",
        "mpp", "mpp__mcc", "mpp__mcc__bmc", "accepted_by",
    )
    for token in tokens:
        qs = qs.filter(
            Q(member_name__icontains=token)       | Q(father_name__icontains=token)  |
            Q(form_number__icontains=token)        | Q(unique_member_code__icontains=token) |
            Q(member_tr_code__icontains=token)     | Q(member_ex_code__icontains=token) |
            Q(mobile_no__icontains=token)          | Q(phone_no__icontains=token)    |
            Q(aadhar_no__icontains=token)          | Q(bank_account_no__icontains=token) |
            Q(ifsc__icontains=token)               | Q(member_bank_name__icontains=token) |
            Q(member_branch_name__icontains=token) | Q(pincode__icontains=token)    |
            Q(village__name__icontains=token)      | Q(tehsil__name__icontains=token) |
            Q(district__name__icontains=token)     | Q(state__name__icontains=token) |
            Q(mpp__name__icontains=token)          | Q(mpp__unique_code__icontains=token) |
            Q(mpp__transaction_code__icontains=token) | Q(mpp__ex_code__icontains=token) |
            Q(mpp__route_name__icontains=token)    | Q(mpp__mcc__name__icontains=token) |
            Q(mpp__mcc__bmc__name__icontains=token)
        )
    results = []
    for f in qs.distinct()[:limit]:
        results.append({
            "id": f.pk, "form_number": f.form_number,
            "unique_member_code": f.unique_member_code,
            "member_tr_code": f.member_tr_code or "",
            "member_ex_code": f.member_ex_code or "",
            "member_name": f.member_name, "father_name": f.father_name or "",
            "member_relation": f.member_relation or "",
            "gender": f.gender, "age": f.age or "",
            "birth_date": str(f.birth_date) if f.birth_date else "",
            "caste_category": f.caste_category or "",
            "qualification": f.qualification or "",
            "aadhar_no": f.aadhar_no or "", "mobile_no": f.mobile_no or "",
            "phone_no": f.phone_no or "", "house_no": f.house_no or "",
            "village": f.village.name  if f.village  else "",
            "village_code": f.village.code  if f.village  else "",
            "tehsil":  f.tehsil.name   if f.tehsil   else "",
            "tehsil_code": f.tehsil.code   if f.tehsil   else "",
            "district": f.district.name if f.district else "",
            "district_code": f.district.code if f.district else "",
            "state": f.state.name    if f.state    else "",
            "pincode": f.pincode or "",
            "mpp_unique_code": f.mpp.unique_code      if f.mpp else "",
            "mpp_name":        f.mpp.name             if f.mpp else "",
            "mpp_tr_code":     f.mpp.transaction_code if f.mpp else "",
            "mcc_name": f.mpp.mcc.name           if f.mpp and f.mpp.mcc else "",
            "bmc_name": f.mpp.mcc.bmc.name       if f.mpp and f.mpp.mcc and f.mpp.mcc.bmc else "",
            "cow_heifer": f.cow_heifer_no,   "buffalo_heifer": f.buffalo_heifer_no,
            "mix_heifer": f.mix_heifer_no,   "cow_dry": f.cow_dry_no,
            "buffalo_dry": f.buffalo_dry_no, "mix_dry": f.mix_dry_no,
            "cow_total": f.cow_animal_nos,   "buffalo_total": f.buffalo_animal_nos,
            "mix_total": f.mix_animal_nos,
            "bank_account_no": f.bank_account_no or "",
            "member_bank_name": f.member_bank_name or "",
            "member_branch_name": f.member_branch_name or "",
            "ifsc": f.ifsc or "",
            "nominee_name": f.nominee_name or "",
            "nominee_relation": f.nominee_relation or "",
            "member_type": f.member_type, "member_status": f.member_status,
            "approval_status": f.approval_status,
            "accepted_by": _safe_accepted_by(f),
            "approval_date": str(f.approval_date)   if f.approval_date   else "",
            "enrollment_date": str(f.enrollment_date) if f.enrollment_date else "",
            "ticket_count": f.tickets.count(),
        })
    return _json_ok(results=results)


# ─────────────────────────────────────────────────────────────────────────────
#  SAHAYAK (MPP) SEARCH
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def sahayak_search(request):
    raw   = request.GET.get("q", "").strip()
    limit = min(int(request.GET.get("limit", 30)), 200)
    if not raw:
        return _json_ok(results=[])
    tokens = [t for t in raw.split() if len(t) >= 2]
    if not tokens:
        return _json_ok(results=[])

    qs = MPP.objects.select_related(
        "village", "tehsil", "district", "state",
        "mcc", "mcc__bmc", "mcc__bmc__plant",
    )
    for token in tokens:
        qs = qs.filter(
            Q(name__icontains=token)             | Q(short_name__icontains=token) |
            Q(unique_code__icontains=token)      | Q(transaction_code__icontains=token) |
            Q(ex_code__icontains=token)          | Q(route_name__icontains=token) |
            Q(route_ex_code__icontains=token)    | Q(mobile_number__icontains=token) |
            Q(dpu_station_code__icontains=token) | Q(dpu_vendor_code__icontains=token) |
            Q(pincode__icontains=token)          | Q(village__name__icontains=token) |
            Q(tehsil__name__icontains=token)     | Q(district__name__icontains=token) |
            Q(state__name__icontains=token)      | Q(mcc__name__icontains=token) |
            Q(mcc__code__icontains=token)        | Q(mcc__bmc__name__icontains=token) |
            Q(mcc__bmc__code__icontains=token)   | Q(mcc__bmc__plant__name__icontains=token)
        )
    results = []
    for m in qs.distinct()[:limit]:
        results.append({
            "id": m.pk, "mpp_name": m.name, "mpp_unique_code": m.unique_code,
            "mpp_tr_code": m.transaction_code or "", "mpp_ex_code": m.ex_code or "",
            "short_name": m.short_name or "", "route_name": m.route_name or "",
            "route_ex_code": m.route_ex_code or "", "mobile_number": m.mobile_number or "",
            "dpu_station_code": m.dpu_station_code or "",
            "dpu_vendor_code": m.dpu_vendor_code or "",
            "opening_date": str(m.opening_date) if m.opening_date else "",
            "status": m.status,
            "village":  m.village.name  if m.village  else "",
            "tehsil":   m.tehsil.name   if m.tehsil   else "",
            "district": m.district.name if m.district else "",
            "state":    m.state.name    if m.state    else "",
            "pincode":  m.pincode or "",
            "mcc_name": m.mcc.name           if m.mcc               else "",
            "mcc_code": m.mcc.code           if m.mcc               else "",
            "bmc_name": m.mcc.bmc.name       if m.mcc and m.mcc.bmc else "",
            "bmc_code": m.mcc.bmc.code       if m.mcc and m.mcc.bmc else "",
            "plant_name": m.mcc.bmc.plant.name if m.mcc and m.mcc.bmc and m.mcc.bmc.plant else "",
            "ticket_count": m.tickets.count(),
        })
    return _json_ok(results=results)


# ─────────────────────────────────────────────────────────────────────────────
#  TRANSPORTER SEARCH
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def transporter_search(request):
    raw   = request.GET.get("q", "").strip()
    limit = min(int(request.GET.get("limit", 30)), 200)
    if not raw:
        return _json_ok(results=[])
    tokens = [t for t in raw.split() if len(t) >= 2]
    if not tokens:
        return _json_ok(results=[])

    qs = Transporter.objects.all()
    for token in tokens:
        qs = qs.filter(
            Q(vendor_code__icontains=token)      | Q(vendor_name__icontains=token) |
            Q(contact_person__icontains=token)   | Q(contact_no__icontains=token) |
            Q(email__icontains=token)            | Q(address__icontains=token) |
            Q(city__icontains=token)             | Q(bank_account_no__icontains=token) |
            Q(bank_key__icontains=token)         | Q(account_holder__icontains=token) |
            Q(gst_number__icontains=token)       | Q(msme__icontains=token) |
            Q(gl_account__icontains=token)       | Q(incoterm_location__icontains=token) |
            Q(search_term1__icontains=token)     | Q(search_term2__icontains=token)
        )
    results = []
    for t in qs.distinct()[:limit]:
        results.append({
            "id": t.pk, "vendor_code": t.vendor_code, "vendor_name": t.vendor_name,
            "account_group": t.account_group,
            "contact_person": t.contact_person or "", "contact_no": t.contact_no or "",
            "email": t.email or "", "address": t.address or "", "city": t.city or "",
            "country": t.country or "", "incoterm": t.incoterm or "",
            "incoterm_location": t.incoterm_location or "",
            "bank_account_no": t.bank_account_no or "", "bank_key": t.bank_key or "",
            "account_holder": t.account_holder or "",
            "payment_terms": t.payment_terms or "", "payment_method": t.payment_method or "",
            "gst_number": t.gst_number or "", "msme": t.msme or "",
            "is_blocked": t.is_blocked, "company_code": t.company_code or "",
            "gl_account": t.gl_account or "", "pan": t.pan_from_gst or "",
            "ticket_count": t.tickets.count(),
        })
    return _json_ok(results=results)


# ─────────────────────────────────────────────────────────────────────────────
#  EMPLOYEE SEARCH
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def employee_search(request):
    q     = request.GET.get("q", "").strip()
    limit = min(int(request.GET.get("limit", 20)), 50)
    if not q:
        return _json_ok(results=[])

    qs = (
        CustomUser.objects
        .filter(account_status=CustomUser.AccountStatus.ACTIVE, login_status=True)
        .filter(
            _q_icontains("first_name",    q) | _q_icontains("last_name",     q) |
            _q_icontains("employee_code", q) | _q_icontains("employee_title", q) |
            _q_icontains("department",    q) | _q_icontains("work_address",   q) |
            _q_icontains("work_phone",    q)
        )
        .select_related("manager")[:limit]
    )
    results = []
    for e in qs:
        mgr = e.manager
        results.append({
            "id": e.pk, "emp_id": e.employee_code,
            "fname": e.first_name, "lname": e.last_name,
            "full_name": e.get_full_name(), "email": e.email,
            "title": e.employee_title or "", "dept": e.department,
            "type": e.employee_type, "location": e.work_address or "",
            "work_phone": e.work_phone or "", "home_phone": e.home_phone or "",
            "manager_id":   mgr.pk              if mgr else None,
            "manager_name": mgr.get_full_name() if mgr else "",
            "manager_code": mgr.employee_code   if mgr else "",
            "manager_dept": mgr.department      if mgr else "",
        })
    return _json_ok(results=results)


# ─────────────────────────────────────────────────────────────────────────────
#  TICKET LIST
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def ticket_list(request):
    status   = request.GET.get("status",   "").strip()
    priority = request.GET.get("priority", "").strip()
    entity   = request.GET.get("entity",   "").strip()
    ttype    = request.GET.get("type",     "").strip()
    q        = request.GET.get("q",        "").strip()
    page     = max(1, int(request.GET.get("page", 1)))
    per_page = min(int(request.GET.get("per_page", 10)), 100)

    qs = (
        Ticket.objects
        .select_related(
            "farmer", "farmer__village", "farmer__tehsil",
            "mpp", "mpp__village", "mpp__district",
            "transporter", "resolved_by", "created_by",
        )
        .prefetch_related("assigned_to")
    )
    if status:
        qs = qs.filter(is_escalated=True) if status == "escalated" else qs.filter(status=status)
    if priority: qs = qs.filter(priority=priority)
    if entity:   qs = qs.filter(entity_type=entity)
    if ttype:    qs = qs.filter(ticket_type__icontains=ttype)
    if q:
        qs = qs.filter(
            _q_icontains("ticket_id",                q) |
            _q_icontains("farmer__member_name",      q) |
            _q_icontains("farmer__mobile_no",        q) |
            _q_icontains("mpp__name",                q) |
            _q_icontains("transporter__vendor_name", q) |
            _q_icontains("other_caller_name",        q) |
            _q_icontains("farmer__village__name",    q) |
            _q_icontains("mpp__village__name",       q) |
            _q_icontains("transporter__city",        q)
        )

    total   = qs.count()
    start   = (page - 1) * per_page
    results = []
    for t in qs.order_by("-created_at")[start: start + per_page]:
        assigned = [
            {"id": u.pk, "name": u.get_full_name(), "code": u.employee_code}
            for u in t.assigned_to.all()
        ]
        results.append({
            "id": t.ticket_id, "pk": t.pk, "entity": t.entity_type,
            "name": t.caller_display_name, "location": t.caller_location,
            "mobile": t.caller_contact_mobile or "",
            "type": t.ticket_type, "priority": t.priority, "status": t.status,
            "is_escalated": t.is_escalated, "is_overdue": t.is_overdue,
            "assigned": assigned,
            "assigned_str": ", ".join(a["name"] for a in assigned) if assigned else "",
            "created": t.created_at.strftime("%d %b %Y, %I:%M %p"),
            "expected": str(t.expected_resolution_date) if t.expected_resolution_date else "",
            "resolved": t.resolved_at.strftime("%d %b %Y, %I:%M %p") if t.resolved_at else "",
            "desc_en": t.description_en or "", "desc_hi": t.description_hi or "",
        })
    return _json_ok(results=results, total=total, page=page, per_page=per_page,
                    total_pages=max(1, -(-total // per_page)))


# ─────────────────────────────────────────────────────────────────────────────
#  TICKET ACTIVITY — GET + POST
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
def ticket_activity(request, ticket_id):
    try:
        t = (
            Ticket.objects
            .select_related(
                "farmer", "farmer__village", "farmer__tehsil",
                "mpp", "mpp__village", "mpp__district",
                "transporter", "resolved_by", "created_by",
            )
            .prefetch_related("assigned_to")
            .get(ticket_id=ticket_id)
        )
    except Ticket.DoesNotExist:
        return _json_error("Ticket not found.", status=404)

    if request.method == "POST":
        if t.status == Ticket.Status.CLOSED:
            return _json_error("This ticket is closed. No further comments can be added.")

        body_html  = request.POST.get("body_html",  "").strip()
        body_text  = request.POST.get("body_text",  "").strip()
        body_hindi = request.POST.get("body_hindi", "").strip()
        hindi_fb   = request.POST.get("hindi_fallback", "false").lower() == "true"
        files      = request.FILES.getlist("attachments")

        if not body_text and not files:
            return _json_error("Please write a comment or attach at least one file.")

        comment = TicketComment.objects.create(
            ticket=t, body_html=body_html or body_text, body_text=body_text,
            body_hindi=body_hindi or None, hindi_fallback=hindi_fb,
            posted_by=request.user, is_internal=False,
        )
        for uploaded in files:
            file_name  = uploaded.name
            file_type  = _file_type_from_name(file_name)
            mime_type  = uploaded.content_type or (mimetypes.guess_type(file_name)[0] or "")
            file_bytes = uploaded.read()
            att = TicketCommentAttachment(
                comment=comment, file_name=file_name, file_type=file_type,
                file_size_bytes=len(file_bytes), mime_type=mime_type,
                uploaded_by=request.user,
            )
            att.file.save(file_name, ContentFile(file_bytes), save=True)

        act_type = (
            TicketActivity.ActivityType.ATTACHMENT if files and not body_text
            else TicketActivity.ActivityType.COMMENT
        )
        TicketActivity.objects.create(
            ticket=t, activity_type=act_type, performed_by=request.user,
            description=(
                f"Comment by {request.user.get_full_name()}"
                + (f" with {len(files)} file(s)" if files else "")
            ),
            comment=comment,
        )
        events = _serialize_activities(_get_activities_qs(t))
        return _json_ok(activities=events, events=events)

    assigned = [
        {"id": u.pk, "name": u.get_full_name(), "code": u.employee_code}
        for u in t.assigned_to.all()
    ]
    ticket_data = {
        "id": t.ticket_id, "pk": t.pk, "entity": t.entity_type,
        "name": t.caller_display_name, "location": t.caller_location,
        "mobile": t.caller_contact_mobile or "",
        "type": t.ticket_type, "priority": t.priority, "status": t.status,
        "is_escalated": t.is_escalated, "is_overdue": t.is_overdue,
        "assigned": assigned, "assigned_str": ", ".join(a["name"] for a in assigned),
        "created": t.created_at.strftime("%d %b %Y, %I:%M %p"),
        "desc_en": t.description_en or "", "desc_hi": t.description_hi or "",
    }
    events = _serialize_activities(_get_activities_qs(t))
    return _json_ok(ticket=ticket_data, events=events, activities=events)


# ─────────────────────────────────────────────────────────────────────────────
#  TICKET SEARCH
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def ticket_search(request):
    q     = request.GET.get("q", "").strip()
    limit = min(int(request.GET.get("limit", 8)), 30)
    if not q:
        return _json_ok(results=[])

    qs = (
        Ticket.objects
        .select_related("farmer", "mpp", "transporter")
        .prefetch_related("assigned_to")
        .filter(
            _q_icontains("ticket_id",                q) |
            _q_icontains("farmer__member_name",      q) |
            _q_icontains("mpp__name",                q) |
            _q_icontains("transporter__vendor_name", q) |
            _q_icontains("other_caller_name",        q) |
            _q_icontains("farmer__village__name",    q) |
            _q_icontains("transporter__city",        q)
        )
        .order_by("-created_at")[:limit]
    )
    return _json_ok(results=[
        {"id": t.ticket_id, "name": t.caller_display_name,
         "location": t.caller_location, "type": t.ticket_type,
         "status": t.status, "priority": t.priority}
        for t in qs
    ])


# ─────────────────────────────────────────────────────────────────────────────
#  TICKET CREATE
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
def ticket_create(request):
    if request.content_type and "multipart/form-data" in request.content_type:
        data = request.POST.copy()
        files = request.FILES.getlist("attachments")
        entity_type     = data.get("entity_type", "other")
        ticket_type     = data.get("ticket_type", "Others")
        priority        = data.get("priority", "medium")
        desc_en         = data.get("description_en", "").strip()
        desc_hi         = data.get("description_hi", "").strip()
        expected        = data.get("expected_resolution_date") or None
        farmer_pk       = data.get("farmer_pk")
        mpp_pk          = data.get("mpp_pk")
        transporter_pk  = data.get("transporter_pk")
        caller_name     = data.get("caller_name", "")
        caller_mobile   = data.get("caller_mobile", "")
        caller_relation = data.get("caller_relation", "")
        on_behalf_of    = data.get("on_behalf_of", "")
        other_name      = data.get("other_caller_name", "")
        other_mobile    = data.get("other_caller_mobile", "")
        other_location  = data.get("other_caller_location", "")
        assignee_pks    = []
        raw_ids = data.get("assignee_ids", "")
        if raw_ids:
            try:
                assignee_pks = json.loads(raw_ids)
            except json.JSONDecodeError:
                assignee_pks = [int(x) for x in raw_ids.split(",") if x]
    else:
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return _json_error("Invalid JSON body.")
        entity_type     = body.get("entity_type", "other")
        ticket_type     = body.get("ticket_type", "Others")
        priority        = body.get("priority", "medium")
        desc_en         = body.get("description_en", "")
        desc_hi         = body.get("description_hi", "")
        assignee_pks    = body.get("assignee_ids", [])
        expected        = body.get("expected_resolution_date") or None
        farmer_pk       = body.get("farmer_pk")
        mpp_pk          = body.get("mpp_pk")
        transporter_pk  = body.get("transporter_pk")
        caller_name     = body.get("caller_name", "")
        caller_mobile   = body.get("caller_mobile", "")
        caller_relation = body.get("caller_relation", "")
        on_behalf_of    = body.get("on_behalf_of", "")
        other_name      = body.get("other_caller_name", "")
        other_mobile    = body.get("other_caller_mobile", "")
        other_location  = body.get("other_caller_location", "")
        files           = []

    farmer_obj      = Farmer.objects.select_related("mpp").filter(pk=farmer_pk).first() if farmer_pk else None
    mpp_obj         = MPP.objects.filter(pk=mpp_pk).first() if mpp_pk else None
    transporter_obj = Transporter.objects.filter(pk=transporter_pk).first() if transporter_pk else None
    if entity_type == "farmer" and farmer_obj and not mpp_obj:
        mpp_obj = farmer_obj.mpp

    exp_date = None
    if expected:
        try:
            exp_date = datetime.datetime.strptime(expected, "%Y-%m-%d").date()
        except Exception:
            pass

    ticket = Ticket.objects.create(
        entity_type=entity_type, farmer=farmer_obj, mpp=mpp_obj,
        transporter=transporter_obj, ticket_type=ticket_type, priority=priority,
        description_en=desc_en, description_hi=desc_hi,
        expected_resolution_date=exp_date, created_by=request.user,
        caller_name=caller_name, caller_mobile=caller_mobile,
        caller_relation=caller_relation, on_behalf_of=on_behalf_of,
        other_caller_name=other_name, other_caller_mobile=other_mobile,
        other_caller_location=other_location,
    )

    assignees = []
    if assignee_pks:
        assignees = list(CustomUser.objects.filter(pk__in=assignee_pks))
        if assignees:
            ticket.assigned_to.set(assignees)

    TicketActivity.objects.create(
        ticket=ticket, activity_type=TicketActivity.ActivityType.CREATED,
        performed_by=request.user, description="Ticket created via Caller Dashboard",
        new_status=Ticket.Status.OPEN,
    )
    if assignees:
        act_assign = TicketActivity.objects.create(
            ticket=ticket, activity_type=TicketActivity.ActivityType.ASSIGNED,
            performed_by=request.user,
            description=f"Assigned to {', '.join(u.get_full_name() for u in assignees)}",
        )
        act_assign.assigned_to.set(assignees)
        ticket.refresh_from_db()
        ticket_assigned_notification(ticket=ticket, assignees=assignees,
                                     created_by=request.user, include_attachments=True)
        ticket_manager_area_alert(ticket=ticket, assignees=assignees, created_by=request.user)

    if files:
        for uploaded in files:
            try:
                file_name  = uploaded.name
                file_type  = _file_type_from_name(file_name)
                mime_type  = uploaded.content_type or (mimetypes.guess_type(file_name)[0] or "")
                file_bytes = uploaded.read()
                att = TicketAttachment(
                    ticket=ticket, file_name=file_name, file_type=file_type,
                    file_size_bytes=len(file_bytes), mime_type=mime_type,
                    uploaded_by=request.user,
                )
                att.file.save(file_name, ContentFile(file_bytes), save=True)
                TicketActivity.objects.create(
                    ticket=ticket, activity_type=TicketActivity.ActivityType.ATTACHMENT,
                    performed_by=request.user,
                    description=f"File attached: {file_name} ({_fmt_size(len(file_bytes))})",
                )
            except Exception as e:
                logger.error("Failed to save attachment %s: %s", uploaded.name, e)

    resp = {"ticket_id": ticket.ticket_id, "pk": ticket.pk}
    if files:
        resp["attachments_count"] = len(files)
        resp["attachments"] = [
            {"name": a.file_name, "size": a.file_size_bytes, "size_display": _fmt_size(a.file_size_bytes)}
            for a in ticket.attachments.all()
        ]
    return _json_ok(**resp)


# ─────────────────────────────────────────────────────────────────────────────
#  TICKET STATUS MUTATIONS
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_POST
def ticket_resolve(request, ticket_id):
    t, err = _get_ticket_or_error(ticket_id)
    if err: return err
    if t.status in (Ticket.Status.RESOLVED, Ticket.Status.CLOSED):
        return _json_error("Ticket is already resolved or closed.")
    old = t.status
    t.mark_resolved(request.user)
    TicketActivity.objects.create(
        ticket=t, activity_type=TicketActivity.ActivityType.RESOLVED,
        performed_by=request.user, old_status=old, new_status=Ticket.Status.RESOLVED,
        description="Ticket resolved via Caller Dashboard",
    )
    ticket_status_notification(t, action="resolved", performed_by=request.user)
    return _json_ok(status=t.status)


@login_required(login_url="login")
@require_POST
def ticket_close(request, ticket_id):
    t, err = _get_ticket_or_error(ticket_id)
    if err: return err
    if t.status == Ticket.Status.CLOSED:
        return _json_error("Ticket is already closed.")
    old = t.status
    t.mark_closed(request.user)
    TicketActivity.objects.create(
        ticket=t, activity_type=TicketActivity.ActivityType.CLOSED,
        performed_by=request.user, old_status=old, new_status=Ticket.Status.CLOSED,
        description="Ticket closed via Caller Dashboard",
    )
    ticket_status_notification(t, action="closed", performed_by=request.user)
    return _json_ok(status=t.status)


@login_required(login_url="login")
@require_POST
def ticket_reopen(request, ticket_id):
    t, err = _get_ticket_or_error(ticket_id)
    if err: return err
    if t.status in (Ticket.Status.OPEN, Ticket.Status.REOPENED):
        return _json_error("Ticket is already open or reopened.")
    old = t.status
    t.reopen()
    TicketActivity.objects.create(
        ticket=t, activity_type=TicketActivity.ActivityType.REOPENED,
        performed_by=request.user, old_status=old, new_status=Ticket.Status.REOPENED,
        description="Ticket reopened via Caller Dashboard",
    )
    ticket_status_notification(t, action="reopened", performed_by=request.user)
    return _json_ok(status=t.status)


@login_required(login_url="login")
@require_POST
def ticket_escalate(request, ticket_id):
    t, err = _get_ticket_or_error(ticket_id)
    if err: return err
    if t.is_escalated:
        return _json_error("Ticket is already escalated.")
    try:
        body = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        body = {}
    reason = body.get("reason", "").strip()
    old = t.status
    t.escalate(escalated_to_user=request.user, reason=reason)
    TicketActivity.objects.create(
        ticket=t, activity_type=TicketActivity.ActivityType.ESCALATED,
        performed_by=request.user, old_status=old, new_status=Ticket.Status.ESCALATED,
        description=f"Ticket escalated{': ' + reason if reason else ''}",
    )
    ticket_escalation_notification(ticket=t, escalated_by=request.user, reason=reason)

    # ── ESCALATION ENGINE: fire Tier-1 immediately ────────────────────────────
    from .tasks import send_tier1_escalation, send_tier2_escalation
    from django.conf import settings
    from datetime import timedelta

    send_tier1_escalation.delay(t.ticket_id)

    # Schedule Tier-2 TIER2_DELAY_HOURS later as a safety net
    # (the periodic sweep is also a fallback, but explicit scheduling is more reliable)
    delay_hours = getattr(settings, "ESCALATION_TIER2_DELAY_HOURS", 24)
    send_tier2_escalation.apply_async(
        args=[t.ticket_id],
        countdown=delay_hours * 3600,
    )

    return _json_ok(status=t.status, is_escalated=t.is_escalated)
 



@login_required(login_url="login")
@require_POST
def ticket_update(request, ticket_id):
    t, err = _get_ticket_or_error(ticket_id)
    if err: return err
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_error("Invalid JSON body.")
    changes = []
    new_priority = None
    if "ticket_type" in body and body["ticket_type"] != t.ticket_type:
        changes.append(f"Type → {body['ticket_type']}")
        t.ticket_type = body["ticket_type"]
    if "priority" in body and body["priority"] != t.priority:
        old_p, new_p = t.priority, body["priority"]
        changes.append(f"Priority: {old_p} → {new_p}")
        t.priority = new_p
        new_priority = new_p
    if "description_en" in body:
        t.description_en = body["description_en"]
    if "caller_mobile" in body:
        t.caller_mobile = body["caller_mobile"]
    t.save()
    if changes:
        TicketActivity.objects.create(
            ticket=t, activity_type=TicketActivity.ActivityType.STATUS_CHANGE,
            performed_by=request.user,
            description="Edited: " + ", ".join(changes), new_priority=new_priority,
        )
    return _json_ok(ticket_id=t.ticket_id)


@login_required(login_url="login")
@require_POST
def ticket_reassign(request, ticket_id):
    t, err = _get_ticket_or_error(ticket_id)
    if err: return err
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return _json_error("Invalid JSON body.")
    assignee_pks = body.get("assignee_ids", [])
    if not assignee_pks:
        return _json_error("Please provide at least one assignee.")
    assignees = list(CustomUser.objects.filter(pk__in=assignee_pks))
    if not assignees:
        return _json_error("No valid employees found for provided IDs.")
    old_assignees = list(t.assigned_to.all())
    t.assigned_to.set(assignees)
    t.save(update_fields=["updated_at"])
    act = TicketActivity.objects.create(
        ticket=t, activity_type=TicketActivity.ActivityType.REASSIGNED,
        performed_by=request.user,
        description=f"Reassigned to {', '.join(u.get_full_name() for u in assignees)}",
    )
    act.assigned_to.set(assignees)
    ticket_reassign_notification(
        ticket=t, new_assignees=assignees,
        old_assignees=old_assignees, reassigned_by=request.user,
    )
    return _json_ok(
        ticket_id=t.ticket_id,
        assigned=[{"id": u.pk, "name": u.get_full_name()} for u in assignees],
    )

# ─────────────────────────────────────────────────────────────────────────────
#  SAHAYAK PAST TICKETS
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def sahayak_tickets(request, mpp_pk):
    try:
        mpp = MPP.objects.get(pk=mpp_pk)
    except MPP.DoesNotExist:
        return _json_error("MPP not found.", status=404)

    qs = (
        Ticket.objects
        .filter(mpp=mpp)
        .select_related("created_by", "resolved_by")
        .prefetch_related("assigned_to")
        .order_by("-created_at")
    )

    results = []
    for t in qs:
        assigned = [
            {"id": u.pk, "name": u.get_full_name(), "code": u.employee_code}
            for u in t.assigned_to.all()
        ]
        results.append({
            "id":           t.ticket_id,
            "pk":           t.pk,
            "type":         t.ticket_type,
            "priority":     t.priority,
            "status":       t.status,
            "is_escalated": t.is_escalated,
            "is_overdue":   t.is_overdue,
            "assigned":     assigned,
            "assigned_str": ", ".join(a["name"] for a in assigned) if assigned else "",
            "created":      t.created_at.strftime("%d %b %Y, %I:%M %p"),
            "resolved":     t.resolved_at.strftime("%d %b %Y, %I:%M %p") if t.resolved_at else "",
        })

    return _json_ok(
        results=results,
        total=len(results),
        mpp_name=mpp.name,
        mpp_code=mpp.unique_code or mpp.transaction_code or "",
    )


# ─────────────────────────────────────────────────────────────────────────────
#  TRANSPORTER PAST TICKETS
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def transporter_tickets(request, trans_pk):
    try:
        transporter = Transporter.objects.get(pk=trans_pk)
    except Transporter.DoesNotExist:
        return _json_error("Transporter not found.", status=404)

    qs = (
        Ticket.objects
        .filter(transporter=transporter)
        .select_related("created_by", "resolved_by")
        .prefetch_related("assigned_to")
        .order_by("-created_at")
    )

    results = []
    for t in qs:
        assigned = [
            {"id": u.pk, "name": u.get_full_name(), "code": u.employee_code}
            for u in t.assigned_to.all()
        ]
        results.append({
            "id":           t.ticket_id,
            "pk":           t.pk,
            "type":         t.ticket_type,
            "priority":     t.priority,
            "status":       t.status,
            "is_escalated": t.is_escalated,
            "is_overdue":   t.is_overdue,
            "assigned":     assigned,
            "assigned_str": ", ".join(a["name"] for a in assigned) if assigned else "",
            "created":      t.created_at.strftime("%d %b %Y, %I:%M %p"),
            "resolved":     t.resolved_at.strftime("%d %b %Y, %I:%M %p") if t.resolved_at else "",
        })

    return _json_ok(
        results=results,
        total=len(results),
        vendor_name=transporter.vendor_name,
        vendor_code=transporter.vendor_code,
    )

# ─────────────────────────────────────────────────────────────────────────────
#  FARMER TICKET HISTORY
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def farmer_tickets(request, farmer_pk):
    try:
        farmer = Farmer.objects.get(pk=farmer_pk)
    except Farmer.DoesNotExist:
        return _json_error("Farmer not found.", status=404)

    qs = (
        Ticket.objects.filter(farmer=farmer)
        .select_related("farmer", "farmer__village", "farmer__tehsil",
                        "mpp", "resolved_by", "created_by")
        .prefetch_related("assigned_to")
        .order_by("-created_at")
    )
    results = []
    for t in qs:
        assigned = [
            {"id": u.pk, "name": u.get_full_name(), "code": u.employee_code}
            for u in t.assigned_to.all()
        ]
        results.append({
            "id": t.ticket_id, "pk": t.pk,
            "type": t.ticket_type, "priority": t.priority, "status": t.status,
            "is_escalated": t.is_escalated, "is_overdue": t.is_overdue,
            "assigned": assigned,
            "assigned_str": ", ".join(a["name"] for a in assigned),
            "created": t.created_at.strftime("%d %b %Y, %I:%M %p"),
            "resolved": t.resolved_at.strftime("%d %b %Y, %I:%M %p") if t.resolved_at else "",
            "desc_en": t.description_en or "",
            "mpp_at_time": t.mpp.name        if t.mpp else (farmer.mpp.name        if farmer.mpp else "—"),
            "mpp_code":    t.mpp.transaction_code if t.mpp else (farmer.mpp.transaction_code if farmer.mpp else "—"),
            "mpp_is_current": not bool(t.mpp),
        })
    return _json_ok(
        results=results, total=len(results),
        farmer_name=farmer.member_name,
        farmer_code=farmer.unique_member_code,
        form_number=farmer.form_number,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT — shared palette + style helpers
# ─────────────────────────────────────────────────────────────────────────────

_G_DARK  = "3A6B4A"; _G_MID   = "4E7D5E"; _G_LIGHT = "C5D9C0"
_NAVY    = "2C3547"; _SLATE   = "4A5568"; _ALT     = "F2F6F1"
_WHITE   = "FFFFFF"; _GRID    = "D4DDD0"; _CREAM   = "FFF8DC"; _GOLD = "7B5700"

_STATUS_C = {
    "open":      ("DBEAFE", "1E40AF"), "pending":   ("FEF9C3", "854D0E"),
    "resolved":  ("DCFCE7", "166534"), "closed":    ("F1F5F9", "475569"),
    "escalated": ("FEE2E2", "991B1B"), "reopened":  ("F3EBF9", "7B3FA0"),
}
_PRIO_C = {
    "low":      ("DCFCE7", "166534"), "medium":   ("FEF9C3", "854D0E"),
    "high":     ("FFEDD5", "9A3412"), "critical": ("FEE2E2", "991B1B"),
}
_ACT_DOT = {
    "created": "2C3547", "assigned": "0A7A8E", "reassigned": "435663",
    "comment": "6F42C1", "attachment": "435663", "resolved": "1E7E34",
    "closed": "5A6270", "reopened": "FD7E14", "escalated": "C0392B",
    "status_change": "B8860B", "priority_change": "B8860B",
    "sms_sent": "138496", "pending": "B8860B",
}
_ACT_LBL = {
    "created": "✦ Created",   "assigned": "→ Assigned",   "reassigned": "⇄ Reassigned",
    "comment": "💬 Comment",  "attachment": "📎 Attachment", "resolved": "✔ Resolved",
    "closed": "🔒 Closed",    "reopened": "↺ Reopened",   "escalated": "⚠ Escalated",
    "status_change": "⊙ Status", "priority_change": "⊙ Priority",
    "sms_sent": "✉ SMS",       "pending": "⏸ Pending",
}
_SLA_OK  = ("1B5E20", "E8F5E9")
_SLA_BAD = ("B71C1C", "FFEBEE")


def _xs(c=_GRID):  return Side(style="thin", color=c)
def _xb():         x = _xs(); return Border(left=x, right=x, top=x, bottom=x)
def _xf(h):        return PatternFill("solid", fgColor=h)
def _xn(bold=False, color=_NAVY, size=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def _xa(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _strip(t): return re.sub(r"<[^>]+>", "", t or "")


# ─────────────────────────────────────────────────────────────────────────────
#  IMAGE HELPERS  (fixed — all three bugs resolved)
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_img_bytes(url, timeout=8):
    """
    Download an image from *url*, resize it to fit a cell, and return
    (BytesIO_at_position_0, pixel_width, pixel_height).

    Returns None on any failure so callers can fall back to text chips.

    Key fixes vs the old helpers:
      • buf.seek(0) is called AFTER save() so openpyxl reads from the start.
      • Resizing happens here — no separate _resize_img / _to_png needed.
      • RGBA conversion prevents palette-PNG transparency errors in openpyxl.
    """
    if not _PIL_OK or not url:
        return None
    try:
        req = _urlreq.Request(
            url,
            headers={"User-Agent": "Mozilla/5.0 (KSTS-Export/1.0)"},
        )
        with _urlreq.urlopen(req, timeout=timeout) as r:
            raw = r.read()

        img = PILImage.open(io.BytesIO(raw))
        img.load()

        # Resize to fit the attachment column
        max_w, max_h = 180, 130
        ow, oh = img.size
        ratio  = min(max_w / ow, max_h / oh, 1.0)
        nw     = max(1, int(ow * ratio))
        nh     = max(1, int(oh * ratio))
        img    = img.resize((nw, nh), PILImage.LANCZOS)

        # Always emit PNG; RGBA avoids palette-transparency issues in openpyxl
        buf = io.BytesIO()
        img.convert("RGBA").save(buf, format="PNG")
        buf.seek(0)   # ← CRITICAL: openpyxl reads from current position

        return buf, nw, nh

    except Exception as exc:
        logger.debug("Image fetch failed for %s: %s", url, exc)
        return None


# ── Column definition for the unified export sheet ───────────────────────────
_TICKET_COLS = [
    ("#",                               "A",  4,   "center"),
    ("Ticket ID",                       "B",  15,  "center"),
    ("Entity",                          "C",  11,  "center"),
    ("Caller",                          "D",  20,  "left"),
    ("Location",                        "E",  18,  "left"),
    ("Mobile",                          "F",  13,  "center"),
    ("Type",                            "G",  20,  "left"),
    ("Priority",                        "H",  11,  "center"),
    ("Status",                          "I",  11,  "center"),
    ("Escalated",                       "J",   9,  "center"),
    ("Assigned To",                     "K",  22,  "left"),
    ("Created By",                      "L",  18,  "left"),
    ("Created At",                      "M",  18,  "center"),
    ("Expected",                        "N",  13,  "center"),
    ("Resolved At",                     "O",  18,  "center"),
    ("SLA Status",                      "P",  22,  "center"),
    ("SLA Hrs",                         "Q",   9,  "center"),
    ("Description (EN)",                "R",  34,  "left"),
    ("Description (हिंदी) / Attachments", "S", 28, "left"),  # wider for images
]
_NCOLS = len(_TICKET_COLS)  # 19


def _write_banner(ws, title, subtitle=""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NCOLS)
    c = ws["A1"]
    c.value     = title
    c.font      = _xn(True, _WHITE, 13)
    c.fill      = _xf(_G_DARK)
    c.alignment = _xa("center")
    ws.row_dimensions[1].height = 26

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_NCOLS)
        c           = ws["A2"]
        c.value     = subtitle
        c.font      = _xn(False, _WHITE, 8, italic=True)
        c.fill      = _xf(_G_MID)
        c.alignment = _xa("left")
        ws.row_dimensions[2].height = 14
        return 3
    return 2


def _write_sla_legend(ws, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
    c           = ws.cell(row=row, column=1)
    c.value     = (
        "SLA Thresholds:   Critical = 1 day (24 hrs)   |   "
        "High = 30 hrs   |   Medium = 48 hrs   |   Low = 72 hrs"
    )
    c.font      = _xn(False, _GOLD, 8, italic=True)
    c.fill      = _xf(_CREAM)
    c.alignment = _xa("center")
    ws.row_dimensions[row].height = 13
    return row + 1


def _write_ticket_header(ws, row):
    for idx, (lbl, col, w, _) in enumerate(_TICKET_COLS, 1):
        c           = ws.cell(row=row, column=idx)
        c.value     = lbl
        c.font      = _xn(True, _NAVY, 9)
        c.fill      = _xf(_G_LIGHT)
        c.alignment = _xa("center")
        c.border    = _xb()
        ws.column_dimensions[col].width = w
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_ticket_row(ws, row, num, t, sla_label, sla_breached, desc_hi=""):
    bg = _WHITE if num % 2 == 1 else _ALT
    sk = (t.get("status")   or "").lower()
    pk = (t.get("priority") or "").lower()
    sb, sf = _STATUS_C.get(sk, ("F1F5F9", "475569"))
    pb, pf = _PRIO_C.get(pk,   ("F1F5F9", "475569"))
    slabg, slafg = _SLA_BAD if sla_breached else _SLA_OK
    esc = t.get("is_escalated", False)

    values = [
        num,
        t.get("ticket_id", ""),
        (t.get("entity_type") or "").capitalize(),
        t.get("caller_name", ""),
        t.get("caller_location", ""),
        t.get("caller_mobile", ""),
        t.get("ticket_type", ""),
        (t.get("priority") or "").capitalize(),
        (t.get("status")   or "").capitalize(),
        "Yes" if esc else "No",
        t.get("assigned_to", ""),
        t.get("created_by", ""),
        t.get("created_at", ""),
        t.get("expected_resolution_date", ""),
        t.get("resolved_at", ""),
        sla_label,
        t.get("sla_threshold_hrs", ""),
        t.get("description_en", ""),
        desc_hi,
    ]
    center_cols = {1, 2, 5, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17}

    for idx, val in enumerate(values, 1):
        c           = ws.cell(row=row, column=idx)
        c.value     = val
        c.border    = _xb()
        c.fill      = _xf(bg)
        c.font      = _xn(color=_SLATE)
        c.alignment = _xa(
            h="center" if idx in center_cols else "left",
            wrap=(idx in {4, 5, 11, 12, 18, 19}),
        )

        if idx == 8:    # Priority
            c.fill = _xf(pb); c.font = _xn(True, pf)
        elif idx == 9:  # Status
            c.fill = _xf(sb); c.font = _xn(True, sf)
        elif idx == 10: # Escalated
            if esc: c.fill = _xf("FFE8E8"); c.font = _xn(True, "C0392B")
        elif idx == 16: # SLA
            c.fill = _xf(slabg); c.font = _xn(True, slafg)
        elif idx == 19 and val: # Hindi description
            c.fill = _xf(_CREAM); c.font = _xn(False, _GOLD, 8, italic=True)

    ws.row_dimensions[row].height = 30
    return row + 1


def _write_timeline_section_header(ws, row, ticket_id):
    """Slim coloured divider labelling the timeline block for a ticket."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
    c           = ws.cell(row=row, column=1)
    c.value     = f"  ▸ Activity Timeline — {ticket_id}"
    c.font      = _xn(True, _WHITE, 8)
    c.fill      = _xf(_G_MID)
    c.alignment = _xa("left")
    ws.row_dimensions[row].height = 13
    return row + 1


def _write_timeline_events(ws, start_row, events):
    """
    Write all activity events for one ticket as sub-rows.

    Layout per event
    ─────────────────────────────────────────────────────────────────────
    • Main text row  : col A (dot) | col B (time) | col C (type+actor)
                       | cols D–R merged (description+note)
                       | col S (non-image file chips)

    • Image rows     : one dedicated row per image attachment, placed
                       BELOW the main text row so they never conflict
                       with the D–R merge.  Cols A–R get a light fill
                       only; col S holds the XLImage anchor.

    • Hindi sub-row  : cols D–S merged, cream background.

    Key fixes applied
    ─────────────────────────────────────────────────────────────────────
    1. buf.seek(0) inside _fetch_img_bytes — openpyxl reads from pos 0.
    2. Images go on SEPARATE rows below the merged text row — no merge
       conflict with the D–R span.
    3. Row height for image rows is set AFTER the image width/height is
       known (nh * 0.75 points).
    4. Failed downloads fall back to a text chip on the main row without
       consuming a spare row.
    """
    cur = start_row

    for ev_idx, ev in enumerate(events):
        atype = ev.get("activity_type") or ev.get("type") or "other"
        dot   = _ACT_DOT.get(atype, "435663")
        lbl   = _ACT_LBL.get(atype, atype.replace("_", " ").title())
        actor = ev.get("actor", "System")
        text  = _strip(ev.get("text", ""))
        note  = (ev.get("note") or "").strip()
        hindi = (ev.get("body_hindi") or "").strip()
        atts  = ev.get("attachments") or []
        ttime = ev.get("time", "")

        tl_bg = "F6FAF6" if ev_idx % 2 == 0 else "EEF4EE"

        desc = text
        if note:
            desc += "\n» " + note[:500]

        img_atts  = [a for a in atts if a.get("is_image") and a.get("url")]
        file_atts = [a for a in atts if not (a.get("is_image") and a.get("url"))]
        chip_txt  = "\n".join(
            "📎 " + (a.get("file_name") or "file")
            + ("  [" + a.get("file_size_display", "") + "]"
               if a.get("file_size_display") else "")
            for a in file_atts
        )

        # ── helper: apply common TL cell style ────────────────────────────
        def _tc(row, col, val="", h="left", wrap=False,
                color=_SLATE, size=8, bold=False, italic=False, bg=tl_bg):
            c = ws.cell(row=row, column=col)
            c.value     = val
            c.font      = Font(name="Calibri", size=size, color=color,
                               bold=bold, italic=italic)
            c.fill      = _xf(bg)
            c.alignment = _xa(h=h, wrap=wrap)
            c.border    = _xb()

        # ── Main text row ──────────────────────────────────────────────────
        _tc(cur, 1, "  ●", color=dot, size=7)
        _tc(cur, 2, ttime, h="center", color="7F8C8D", size=7)

        c3           = ws.cell(row=cur, column=3)
        c3.value     = f"{lbl}\n{actor}"
        c3.font      = Font(name="Calibri", size=7, color=dot)
        c3.fill      = _xf(tl_bg)
        c3.alignment = _xa("left", wrap=True)
        c3.border    = _xb()

        # Cols D..R — description + note (merged)
        ws.merge_cells(start_row=cur, start_column=4, end_row=cur, end_column=18)
        _tc(cur, 4, desc, h="left", wrap=True, color=_NAVY, size=8)

        # Col S — non-image file chips
        _tc(cur, 19, chip_txt, h="left", wrap=True, color="1E55A3", size=7)

        desc_lines = max(1, len(desc.split("\n")))
        chip_lines = max(1, len(chip_txt.split("\n"))) if chip_txt else 1
        ws.row_dimensions[cur].height = max(20, max(desc_lines, chip_lines) * 11 + 4)

        main_text_row = cur   # save for fallback chip updates
        cur += 1

        # ── Image rows — one per image, below the main text row ───────────
        for ia in img_atts:
            result = _fetch_img_bytes(ia.get("url", ""))

            if result is not None:
                buf, nw, nh = result

                # Fresh BytesIO copy — buf is already at position 0 from
                # _fetch_img_bytes, but an extra seek() costs nothing and
                # protects against future refactoring.
                img_buf = io.BytesIO(buf.read())
                img_buf.seek(0)

                try:
                    xl_img        = XLImage(img_buf)
                    xl_img.width  = nw
                    xl_img.height = nh

                    # Anchor to col S of the dedicated image row
                    anchor_cell = f"{get_column_letter(19)}{cur}"
                    ws.add_image(xl_img, anchor_cell)

                    # Row height to accommodate the image
                    # (openpyxl row height unit ≈ 0.75 × pixel height)
                    needed_pts = round(nh * 0.75) + 6
                    ws.row_dimensions[cur].height = max(needed_pts, 22)

                    # Style cols A–R of this image row (grid stays intact)
                    for col in range(1, 19):
                        c = ws.cell(row=cur, column=col)
                        c.fill   = _xf(tl_bg)
                        c.border = _xb()

                    # Col S: white background so the image sits cleanly
                    ws.cell(row=cur, column=19).fill   = _xf(_WHITE)
                    ws.cell(row=cur, column=19).border = _xb()

                    cur += 1   # advance only on success

                except Exception as exc:
                    logger.warning(
                        "Failed to embed image %s: %s",
                        ia.get("file_name", "?"), exc,
                    )
                    # Fallback: add chip text to the main row (no row consumed)
                    cell_s = ws.cell(row=main_text_row, column=19)
                    prev   = cell_s.value or ""
                    cell_s.value = (prev + "\n" if prev else "") + "🖼 " + ia.get("file_name", "image")

            else:
                # Download failed — add chip text to main row (no row consumed)
                name   = ia.get("file_name", "image")
                cell_s = ws.cell(row=main_text_row, column=19)
                prev   = cell_s.value or ""
                cell_s.value = (prev + "\n" if prev else "") + "🖼 " + name

        # ── Hindi sub-row ─────────────────────────────────────────────────
        if hindi:
            ws.merge_cells(
                start_row=cur, start_column=4,
                end_row=cur,   end_column=19,
            )
            hc           = ws.cell(row=cur, column=4)
            hc.value     = "हिंदी: " + hindi[:500]
            hc.font      = Font(name="Calibri", size=7, color=_GOLD, italic=True)
            hc.fill      = _xf(_CREAM)
            hc.alignment = _xa("left", wrap=True)
            hc.border    = _xb()
            for col in [1, 2, 3]:
                cx       = ws.cell(row=cur, column=col)
                cx.fill  = _xf(_CREAM)
                cx.border = _xb()
            ws.row_dimensions[cur].height = max(13, len(hindi) // 8 + 12)
            cur += 1

    # ── Thin green divider after all TL rows for this ticket ──────────────
    for col in range(1, _NCOLS + 1):
        c        = ws.cell(row=cur, column=col)
        c.fill   = _xf(_G_LIGHT)
        c.border = _xb()
        c.value  = ""
    ws.row_dimensions[cur].height = 3
    return cur + 1


def _sla_calc(ticket_obj, now):
    SLA_HOURS = {"critical": 24, "high": 30, "medium": 48, "low": 72}
    hours    = SLA_HOURS.get(ticket_obj.priority, 72)
    deadline = ticket_obj.created_at + timedelta(hours=hours)
    end_time = ticket_obj.resolved_at if ticket_obj.resolved_at else now
    if end_time <= deadline:
        taken = round((end_time - ticket_obj.created_at).total_seconds() / 3600, 1)
        return f"Within SLA ({taken} hrs)", False, hours
    else:
        over  = round((end_time - deadline).total_seconds() / 3600, 1)
        return f"Breached ({over} hrs over)", True, hours


def _build_summary_sheet(wb, tickets_data):
    ws = wb.create_sheet(title="Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14

    ws.merge_cells("A1:B1")
    c       = ws["A1"]; c.value = "Export Summary"
    c.font  = _xn(True, _WHITE, 12)
    c.fill  = _xf(_G_DARK)
    c.alignment = _xa("center")
    ws.row_dimensions[1].height = 22

    status_c   = Counter((t.get("status")      or "").lower() for t in tickets_data)
    priority_c = Counter((t.get("priority")     or "").lower() for t in tickets_data)
    entity_c   = Counter((t.get("entity_type")  or "").lower() for t in tickets_data)
    sla_c      = Counter((t.get("sla_status")   or "")         for t in tickets_data)

    cur = 2
    for sec_lbl, counter in [
        ("By Status", status_c), ("By Priority", priority_c),
        ("By Entity", entity_c), ("By SLA Status", sla_c),
    ]:
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=2)
        h       = ws.cell(row=cur, column=1); h.value = sec_lbl
        h.font  = _xn(True, _NAVY, 10)
        h.fill  = _xf(_G_LIGHT)
        h.alignment = _xa("left")
        cur += 1
        for key, cnt in sorted(counter.items()):
            bg = _WHITE if cur % 2 == 0 else _ALT
            if sec_lbl == "By SLA Status":
                if key == "Within SLA":
                    bg = _SLA_OK[1]; lf = cf = _SLA_OK[0]
                elif key == "Breached":
                    bg = _SLA_BAD[1]; lf = cf = _SLA_BAD[0]
                else:
                    lf = cf = _SLATE
            else:
                lf = cf = _SLATE
            for col, val in [(1, key.capitalize()), (2, cnt)]:
                c       = ws.cell(row=cur, column=col); c.value = val
                c.font  = _xn(color=cf, size=9) if col == 1 else _xn(color=lf, size=9)
                c.fill  = _xf(bg); c.border = _xb()
                c.alignment = _xa("center" if col == 2 else "left")
            cur += 1
        cur += 1

    for col, val in [(1, "Total Tickets"), (2, len(tickets_data))]:
        c       = ws.cell(row=cur, column=col); c.value = val
        c.fill  = _xf(_G_DARK); c.font = _xn(True, _WHITE, 10)
        c.border = _xb(); c.alignment = _xa("center")


def _build_sla_sheet(wb, tickets_data):
    ws = wb.create_sheet(title="SLA Report")
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 16), ("B", 16), ("C", 14), ("D", 14), ("E", 20)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    c       = ws["A1"]; c.value = "SLA Compliance Report"
    c.font  = _xn(True, _WHITE, 12)
    c.fill  = _xf(_G_DARK)
    c.alignment = _xa("center")
    ws.row_dimensions[1].height = 22

    for idx, lbl in enumerate(["Priority", "SLA Threshold", "Within SLA", "Breached", "% Compliance"], 1):
        c       = ws.cell(row=2, column=idx); c.value = lbl
        c.font  = _xn(True, _NAVY, 10)
        c.fill  = _xf(_G_LIGHT)
        c.alignment = _xa("center"); c.border = _xb()

    SLA_LBL   = {"critical": "1 day (24 hrs)", "high": "30 hrs", "medium": "48 hrs", "low": "72 hrs"}
    by_prio   = {p: {"w": 0, "b": 0} for p in ["critical", "high", "medium", "low"]}
    for t in tickets_data:
        p = (t.get("priority") or "").lower()
        if p in by_prio:
            if t.get("sla_status") == "Within SLA": by_prio[p]["w"] += 1
            else:                                    by_prio[p]["b"] += 1

    r = totw = totb = 3, 0, 0
    r = 3
    totw = totb = 0
    for prio in ["critical", "high", "medium", "low"]:
        w   = by_prio[prio]["w"]; b = by_prio[prio]["b"]
        tot = w + b; pct = f"{round(w/tot*100, 1)}%" if tot else "0%"
        totw += w; totb += b
        pb, pf = _PRIO_C.get(prio, ("F1F5F9", "475569"))
        rowbg  = _WHITE if r % 2 == 0 else _ALT
        for idx, val, fg, bg in [
            (1, prio.capitalize(), pf,          pb),
            (2, SLA_LBL.get(prio, "—"), _SLATE, rowbg),
            (3, w,   _SLA_OK[0],  _SLA_OK[1]),
            (4, b,   _SLA_BAD[0] if b else _SLATE, _SLA_BAD[1] if b else rowbg),
            (5, pct, _SLA_OK[0] if (w / tot >= 0.5 if tot else True) else _SLA_BAD[0], rowbg),
        ]:
            c       = ws.cell(row=r, column=idx); c.value = val
            c.font  = _xn(True if idx in (1, 3, 4, 5) else False, fg, 9)
            c.fill  = _xf(bg); c.alignment = _xa("center"); c.border = _xb()
        r += 1

    grand = totw + totb
    gp    = f"{round(totw/grand*100, 1)}%" if grand else "0%"
    for idx, val in enumerate(["TOTAL", "—", totw, totb, gp], 1):
        c       = ws.cell(row=r, column=idx); c.value = val
        c.fill  = _xf(_G_DARK); c.font = _xn(True, _WHITE, 10)
        c.border = _xb(); c.alignment = _xa("center")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN WORKBOOK BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _build_workbook(ticket_rows, ticket_objects, generated_by="", filters=None):
    """
    Parameters
    ----------
    ticket_rows    : list[dict]    — serialised ticket data
    ticket_objects : list[Ticket]  — ORM objects, same order as ticket_rows
    generated_by   : str
    filters        : dict
    """
    filters = filters or {}
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Tickets"
    ws.sheet_view.showGridLines = False

    # ── Banner + meta + SLA legend ────────────────────────────────────────────
    subtitle = (
        f"Generated by: {generated_by}   |   "
        f"Date: {datetime.datetime.now().strftime('%d %b %Y, %I:%M %p')}   |   "
        f"Status: {filters.get('status', 'All')}   |   "
        f"Priority: {filters.get('priority', 'All')}   |   "
        f"Entity: {filters.get('entity', 'All')}   |   "
        f"Type: {filters.get('type', 'All')}   |   "
        f"SLA: {filters.get('sla', 'All')}   |   "
        f"Records: {len(ticket_rows)}"
    )
    next_row = _write_banner(ws, "Ticket System — Complete Export Report", subtitle)
    next_row = _write_sla_legend(ws, next_row)

    # ── Column headers ────────────────────────────────────────────────────────
    next_row = _write_ticket_header(ws, next_row)

    # Override col S width here to ensure images are not clipped
    ws.column_dimensions["S"].width = 28

    ws.freeze_panes = ws.cell(row=next_row, column=1)

    # ── Build ticket_id → ORM object lookup ───────────────────────────────────
    obj_map = {obj.ticket_id: obj for obj in ticket_objects}

    # ── One block per ticket: ticket row + timeline rows ─────────────────────
    for num, t in enumerate(ticket_rows, 1):
        tid          = t.get("ticket_id", "")
        obj          = obj_map.get(tid)
        sla_label    = t.get("sla_label", "—")
        sla_breached = t.get("sla_status") == "Breached"
        desc_hi      = (obj.description_hi or "") if obj else ""

        next_row = _write_ticket_row(ws, next_row, num, t, sla_label, sla_breached, desc_hi)

        if obj:
            try:
                events   = _serialize_activities(_get_activities_qs(obj))
                next_row = _write_timeline_section_header(ws, next_row, tid)
                next_row = _write_timeline_events(ws, next_row, events)
            except Exception as exc:
                logger.warning("Timeline skipped for %s: %s", tid, exc)

    _build_summary_sheet(wb, ticket_rows)
    _build_sla_sheet(wb, ticket_rows)
    return wb


# ─────────────────────────────────────────────────────────────────────────────
#  CALLER DASHBOARD EXPORT  — GET /api/tickets/export/
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
@require_GET
def ticket_export_excel(request):
    status   = request.GET.get("status",   "").strip()
    priority = request.GET.get("priority", "").strip()
    entity   = request.GET.get("entity",   "").strip()
    ttype    = request.GET.get("type",     "").strip()
    q        = request.GET.get("q",        "").strip()
    sla      = request.GET.get("sla",      "").strip()

    qs = (
        Ticket.objects
        .select_related(
            "farmer", "farmer__village", "farmer__tehsil",
            "mpp", "mpp__village", "mpp__district",
            "transporter", "resolved_by", "created_by",
        )
        .prefetch_related("assigned_to")
    )
    if status:
        qs = qs.filter(is_escalated=True) if status == "escalated" else qs.filter(status=status)
    if priority: qs = qs.filter(priority=priority)
    if entity:   qs = qs.filter(entity_type=entity)
    if ttype:    qs = qs.filter(ticket_type__icontains=ttype)
    if q:
        qs = qs.filter(
            Q(ticket_id__icontains=q) | Q(farmer__member_name__icontains=q) |
            Q(mpp__name__icontains=q) | Q(transporter__vendor_name__icontains=q) |
            Q(other_caller_name__icontains=q)
        )

    now         = timezone.now()
    ticket_objs = list(qs.order_by("-created_at"))

    ticket_rows = []
    for t in ticket_objs:
        sla_label, sla_breached, hours = _sla_calc(t, now)
        sla_status = "Breached" if sla_breached else "Within SLA"
        if sla == "within"   and sla_status != "Within SLA": continue
        if sla == "breached" and sla_status != "Breached":   continue
        ticket_rows.append({
            "ticket_id":                t.ticket_id,
            "entity_type":              t.entity_type,
            "caller_name":              t.caller_display_name,
            "caller_location":          t.caller_location,
            "caller_mobile":            t.caller_contact_mobile or "",
            "ticket_type":              t.ticket_type,
            "priority":                 t.priority,
            "status":                   t.status,
            "is_escalated":             t.is_escalated,
            "assigned_to":              ", ".join(u.get_full_name() for u in t.assigned_to.all()),
            "created_by":               t.created_by.get_full_name() if t.created_by else "System",
            "created_at":               t.created_at.strftime("%d %b %Y, %I:%M %p"),
            "expected_resolution_date": str(t.expected_resolution_date) if t.expected_resolution_date else "",
            "resolved_at":              t.resolved_at.strftime("%d %b %Y, %I:%M %p") if t.resolved_at else "",
            "description_en":           t.description_en or "",
            "sla_threshold_hrs":        hours,
            "sla_label":                sla_label,
            "sla_status":               sla_status,
        })

    included    = {r["ticket_id"] for r in ticket_rows}
    ticket_objs = [obj for obj in ticket_objs if obj.ticket_id in included]

    wb = _build_workbook(
        ticket_rows, ticket_objs,
        generated_by=request.user.get_full_name() or request.user.email,
        filters={
            "status":   status   or "All",
            "priority": priority or "All",
            "entity":   entity   or "All",
            "type":     ttype    or "All",
            "sla":      {"within": "Within SLA Only", "breached": "Breached Only"}.get(sla, "All"),
        },
    )

    out   = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"KSTS_Tickets_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp  = HttpResponse(
        out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
#  MY TICKETS EXPORT  — GET /api/my-tickets/export/
# ─────────────────────────────────────────────────────────────────────────────

_MY_SLA_HOURS = {"critical": 24, "high": 30, "medium": 48, "low": 72}


def _my_sla(ticket):
    sla_h   = _MY_SLA_HOURS.get(ticket.priority, 48)
    created = ticket.created_at
    if not created:
        return "N/A", False
    if created.tzinfo is None:
        created = created.replace(tzinfo=datetime.timezone.utc)
    end = ticket.resolved_at or datetime.datetime.now(datetime.timezone.utc)
    if end.tzinfo is None:
        end = end.replace(tzinfo=datetime.timezone.utc)
    elapsed = (end - created).total_seconds() / 3600
    diff    = round(elapsed - sla_h, 1)
    if diff <= 0:
        return f"Within SLA ({abs(diff)} hrs)", False
    return f"Breached ({diff} hrs over)", True


@login_required
@require_GET
def my_tickets_export(request):
    qs = (
        request.user.assigned_tickets.all()
        .select_related(
            "farmer", "farmer__village", "farmer__tehsil",
            "mpp", "mpp__village", "mpp__district",
            "transporter", "created_by",
        )
        .prefetch_related("assigned_to")
        .order_by("-created_at")
    )
    status_f   = request.GET.get("status",   "").strip()
    priority_f = request.GET.get("priority", "").strip()
    entity_f   = request.GET.get("entity",   "").strip()
    type_f     = request.GET.get("type",     "").strip()
    q_f        = request.GET.get("q",        "").strip()
    if status_f:   qs = qs.filter(status=status_f)
    if priority_f: qs = qs.filter(priority=priority_f)
    if entity_f:   qs = qs.filter(entity_type=entity_f)
    if type_f:     qs = qs.filter(ticket_type=type_f)
    if q_f:
        qs = qs.filter(
            Q(ticket_id__icontains=q_f) | Q(farmer__member_name__icontains=q_f) |
            Q(mpp__name__icontains=q_f) | Q(transporter__vendor_name__icontains=q_f) |
            Q(other_caller_name__icontains=q_f) | Q(ticket_type__icontains=q_f)
        )

    ticket_objs = list(qs)
    ticket_rows = []
    for t in ticket_objs:
        sla_label, sla_breached = _my_sla(t)
        sla_status = "Breached" if sla_breached else "Within SLA"
        hours      = _MY_SLA_HOURS.get(t.priority, 48)

        if t.entity_type == "farmer" and t.farmer:
            caller   = t.farmer.member_name or "—"
            v        = getattr(t.farmer, "village", None)
            teh      = getattr(t.farmer, "tehsil",  None)
            location = (
                f"{getattr(v,'name','')} ({getattr(v,'code','')}), "
                f"{getattr(teh,'name','')} ({getattr(teh,'code','')})"
                if v else "—"
            )
            mobile   = t.farmer.mobile_no or "—"
        elif t.entity_type == "sahayak" and t.mpp:
            caller   = t.mpp.name or "—"
            v        = getattr(t.mpp, "village",  None)
            dist     = getattr(t.mpp, "district", None)
            location = (
                f"{getattr(v,'name','')} ({getattr(v,'code','')}), "
                f"{getattr(dist,'name','')} ({getattr(dist,'code','')})"
                if v else "—"
            )
            mobile   = t.mpp.mobile_number or "—"
        elif t.entity_type == "transporter" and t.transporter:
            caller   = t.transporter.vendor_name or "—"
            location = t.transporter.city or "—"
            mobile   = t.transporter.contact_no or "—"
        else:
            caller   = getattr(t, "other_caller_name",     "") or "Unknown"
            location = getattr(t, "other_caller_location", "") or "—"
            mobile   = getattr(t, "other_caller_mobile",   "") or "—"

        assigned_str = "—"
        try:
            names = [a.get_full_name() for a in t.assigned_to.all() if a.get_full_name()]
            if names: assigned_str = ", ".join(names)
        except Exception:
            pass

        ticket_rows.append({
            "ticket_id":                t.ticket_id,
            "entity_type":              "Sahayak" if t.entity_type in ("sahayak", "mpp") else t.entity_type.capitalize(),
            "caller_name":              caller,
            "caller_location":          location,
            "caller_mobile":            mobile,
            "ticket_type":              t.ticket_type,
            "priority":                 t.priority,
            "status":                   t.status,
            "is_escalated":             getattr(t, "is_escalated", False),
            "assigned_to":              assigned_str,
            "created_by":               t.created_by.get_full_name() if t.created_by else "System",
            "created_at":               t.created_at.strftime("%d %b %Y, %I:%M %p") if t.created_at else "—",
            "expected_resolution_date": t.expected_resolution_date.strftime("%Y-%m-%d") if t.expected_resolution_date else "—",
            "resolved_at":              t.resolved_at.strftime("%d %b %Y, %I:%M %p") if t.resolved_at else "—",
            "description_en":           t.description_en or "—",
            "sla_threshold_hrs":        hours,
            "sla_label":                sla_label,
            "sla_status":               sla_status,
        })

    uname = (
        request.user.get_full_name()
        or getattr(request.user, "employee_code", "")
        or "user"
    ).replace(" ", "_")

    wb = _build_workbook(
        ticket_rows, ticket_objs,
        generated_by=request.user.get_full_name() or request.user.email,
        filters={
            "status":   status_f   or "All",
            "priority": priority_f or "All",
            "entity":   entity_f   or "All",
            "type":     type_f     or "All",
            "sla":      "All",
        },
    )

    out   = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"MyTickets_{uname}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp  = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp